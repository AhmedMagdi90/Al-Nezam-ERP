from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
import logging
import re
from io import BytesIO
from difflib import get_close_matches

from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.conf import settings as django_settings
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.contrib.auth import update_session_auth_hash
from urllib.parse import urlparse, parse_qs, urlencode
from openpyxl import Workbook, load_workbook

from accounts.models import Profile, Role
from accounts.constants import RoleType
from manufacturing.models import BOMOperation, Machine, ProductionStage, SystemSettings, WorkOrder
from manufacturing.runtime_translations import (
    canonicalize_translation_key,
    get_company_translation_map,
    list_translation_entries_for_company,
    remove_company_translation,
    resolve_runtime_translation,
    upsert_company_translation,
)
from manufacturing.template_translation_inventory import (
    get_template_translation_inventory,
    list_template_translation_sources,
)
from manufacturing.access_control import resolve_user_role
from manufacturing.shift_utils import (
    SHIFT_LABELS,
    SHIFT_MODE_CHOICES,
    enabled_shift_keys_for_mode,
    normalize_shift_configuration,
    normalize_shift_mode,
)
from .dashboard import redirect_to_role_home, require_company, user_has_role

logger = logging.getLogger(__name__)

class SettingsDashboardView(LoginRequiredMixin, View):
    DEPARTMENT_SPLIT_RE = re.compile(r"[\n;]+")
    VALID_SCOPES = {"manufacturing", "quality", "maintenance"}
    SCOPE_ALIASES = {"planner": "manufacturing", "production": "manufacturing"}
    SELF_SERVICE_ACTIONS = {"update_profile", "update_password"}
    TRANSLATION_SHEET_NAME = "Translations"
    TRANSLATION_SHEET_HEADERS = [
        "English Source",
        "Arabic Translation",
        "Screen Hint",
        "Template Location",
        "Status",
    ]
    DEFAULT_DEPARTMENTS = {
        "manufacturing": [],
        "quality": [],
        "maintenance": [],
    }
    TRANSLATION_PRESET_GROUPS = [
        {
            "key": "reports",
            "label": "Reports Preview",
            "items": [
                "Actual vs Planned",
                "Actual vs Planned Report",
                "Compare production time and material consumption by Work Order.",
                "Apply Filters",
                "All Work Orders",
                "All Products",
                "Export CSV",
                "JSON",
            ],
        },
    ]

    def _tenant_db_alias(self, request):
        return getattr(request, "tenant_db_alias", "default")

    def _ensure_system_roles(self, db_alias):
        for role_key, _label in Role.ROLE_CHOICES:
            Role.objects.using(db_alias).get_or_create(name=role_key)

    def _get_settings_scope(self, request):
        scope = (request.POST.get("scope") or request.GET.get("scope") or "").strip().lower()
        if not scope:
            ref = request.META.get("HTTP_REFERER", "")
            if ref:
                try:
                    scope = (parse_qs(urlparse(ref).query).get("scope", [""])[0] or "").strip().lower()
                except Exception:
                    scope = ""
        scope = self.SCOPE_ALIASES.get(scope, scope)
        if scope not in self.VALID_SCOPES:
            scope = "manufacturing"
        return scope

    def _profile_scope_values(self, scope):
        if scope == "manufacturing":
            return ["manufacturing", "planner"]
        return [scope]

    def _department_catalog_keys_for_scope(self, scope):
        if scope == "manufacturing":
            return ["manufacturing", "planner"]
        return [scope]

    def _redirect_with_scope(self, request, tab="general", extra_params=None):
        scope = self._get_settings_scope(request)
        params = {"tab": tab, "scope": scope}
        if extra_params:
            params.update({key: value for key, value in extra_params.items() if value is not None})
        return redirect(f"{request.path}?{urlencode(params)}")

    def _allowed_role_names_for_scope(self, scope):
        if scope == "quality":
            return [RoleType.QUALITY.value, RoleType.SUPERVISOR.value, RoleType.WORKER.value]
        if scope == "maintenance":
            return [RoleType.MAINTENANCE.value, RoleType.SUPERVISOR.value, RoleType.WORKER.value]
        if scope == "manufacturing":
            return [RoleType.PLANNER.value, RoleType.SUPERVISOR.value, RoleType.WORKER.value, RoleType.STORE.value]
        return [choice[0] for choice in Role.ROLE_CHOICES]

    def _allowed_role_names_for_inviter(self, user):
        role_name = resolve_user_role(user) or ''

        if role_name == RoleType.PLANNER.value:
            return [RoleType.SUPERVISOR.value, RoleType.WORKER.value]
        return [choice[0] for choice in Role.ROLE_CHOICES]

    def _role_skips_department(self, role_name):
        return role_name in {RoleType.PLANNER.value, RoleType.STORE.value}

    def _normalize_member_phone(self, value):
        raw = str(value or "").strip()
        if not raw:
            return ""
        digits = re.sub(r"\D+", "", raw)
        if not digits:
            return ""
        return f"+{digits}" if raw.startswith("+") else digits

    def _member_contact_seed(self, email="", phone=""):
        email = str(email or "").strip().lower()
        phone = self._normalize_member_phone(phone)
        if email:
            return email.split("@")[0]
        return phone or "user"

    def _build_unique_member_username(self, db_alias, company, email="", phone=""):
        base_username = self._member_contact_seed(email=email, phone=phone)[:150] or "user"
        username = base_username
        counter = 1
        while User.objects.using(db_alias).filter(username=username).exists():
            username = f"{base_username}_{company.id}_{counter}"[:150]
            counter += 1
        return username

    def _can_admin_reset_member_password(self, user):
        return (resolve_user_role(user) or "") == RoleType.ADMIN.value

    def _department_catalog_for_scope(self, settings, scope):
        catalog = settings.department_catalog or {}
        values = []
        for key in self._department_catalog_keys_for_scope(scope):
            candidate = catalog.get(key, [])
            if isinstance(candidate, list):
                values.extend(candidate)
        if not isinstance(values, list):
            return []
        return self._split_department_values(values)

    def _clean_department_label(self, value):
        return " ".join(str(value or "").strip().split())

    def _split_department_values(self, value):
        if value is None:
            return []

        if isinstance(value, (list, tuple, set)):
            raw_values = value
        else:
            raw_values = self.DEPARTMENT_SPLIT_RE.split(str(value))

        cleaned = []
        seen = set()
        for item in raw_values:
            label = self._clean_department_label(item)
            key = label.lower()
            if not label or key in seen:
                continue
            seen.add(key)
            cleaned.append(label)
        return cleaned

    def _join_department_values(self, values):
        return "\n".join(self._split_department_values(values))

    def _live_department_options(self, company, db_alias):
        if not company:
            return []

        live_options = []
        seen = set()

        def add_option(value):
            label = self._clean_department_label(value)
            key = label.lower()
            if not label or key in seen:
                return
            seen.add(key)
            live_options.append(label)

        for value in Machine.objects.using(db_alias).filter(company=company).values_list("category", flat=True):
            add_option(value)

        stage_ids = set(
            stage_id
            for stage_id in WorkOrder.objects.using(db_alias).filter(company=company).values_list("stage_id", flat=True)
            if stage_id
        )
        stage_ids.update(
            stage_id
            for stage_id in WorkOrder.objects.using(db_alias).filter(company=company).values_list("current_stage_id", flat=True)
            if stage_id
        )
        stage_ids.update(
            stage_id
            for stage_id in BOMOperation.objects.using(db_alias).filter(bom__product__company=company).values_list("stage_id", flat=True)
            if stage_id
        )
        stage_ids.update(
            ProductionStage.objects.using(db_alias)
            .filter(machine__company=company)
            .values_list("id", flat=True)
        )

        for stage in ProductionStage.objects.using(db_alias).filter(id__in=stage_ids).order_by("order", "name"):
            add_option(stage.category)

        return live_options

    def _department_options_for_scope(self, settings, scope, company=None, db_alias="default"):
        return self._department_catalog_for_scope(settings, scope)

    def _shift_options(self, settings):
        configured = normalize_shift_configuration(
            settings.shift_configuration,
            shift_mode=getattr(settings, "shift_mode", "3"),
            default_enabled=True,
        )

        options = []
        seen = set()
        for key in ("morning", "afternoon", "night"):
            window = configured.get(key, {})
            if not window.get("enabled"):
                continue
            value = "evening" if key == "afternoon" else str(key or "").strip().lower()
            if not value or value in seen:
                continue
            seen.add(value)
            start = str(window.get("start") or "").strip()
            end = str(window.get("end") or "").strip()
            options.append({
                "value": value,
                "label": SHIFT_LABELS.get(key, str(key).replace("_", " ").strip().title()),
                "window": f"{start} - {end}" if start and end else "",
            })
        return options

    def _translation_entries(self, company, language="ar"):
        try:
            entries = list_translation_entries_for_company(company, language, django_settings.BASE_DIR)
        except OSError:
            logger.exception("Failed to load translation catalog for company=%s language=%s", getattr(company, "id", None), language)
            entries = []

        indexed = {entry["msgid"]: dict(entry) for entry in entries}
        company_overrides = get_company_translation_map(company, language)
        for msgid in self._known_translation_sources(company, language):
            if msgid in indexed:
                continue
            indexed[msgid] = {
                "msgid": msgid,
                "msgstr": company_overrides.get(msgid, ""),
                "origin": "company" if msgid in company_overrides else "missing",
            }

        return [indexed[key] for key in sorted(indexed.keys(), key=str.lower)]

    def _translation_preview_groups(self, company, language="ar"):
        company_overrides = get_company_translation_map(company, language)
        groups = []
        for group in self.TRANSLATION_PRESET_GROUPS:
            items = []
            for msgid in group["items"]:
                items.append(
                    {
                        "msgid": msgid,
                        "current_value": resolve_runtime_translation(company, msgid, language),
                        "origin": "company" if msgid in company_overrides else "fallback",
                    }
                )
            groups.append({"key": group["key"], "label": group["label"], "items": items})
        return groups

    def _known_translation_sources(self, company, language="ar"):
        known_sources = set()
        try:
            entries = list_translation_entries_for_company(company, language, django_settings.BASE_DIR)
            known_sources.update(entry["msgid"] for entry in entries if entry.get("msgid"))
        except OSError:
            logger.exception(
                "Failed to load known translation sources for company=%s language=%s",
                getattr(company, "id", None),
                language,
            )

        for group in self.TRANSLATION_PRESET_GROUPS:
            known_sources.update(group["items"])

        known_sources.update(list_template_translation_sources(django_settings.BASE_DIR))

        return sorted(known_sources, key=str.lower)

    def _translation_feedback(self, request):
        return request.session.pop("translation_feedback", None)

    def _translation_sheet_rows(self, company, language="ar"):
        indexed_entries = {
            entry["msgid"]: dict(entry)
            for entry in self._translation_entries(company, language)
        }
        location_inventory = get_template_translation_inventory(django_settings.BASE_DIR)
        rows = []

        for msgid in self._known_translation_sources(company, language):
            entry = indexed_entries.get(
                msgid,
                {"msgid": msgid, "msgstr": "", "origin": "missing"},
            )
            location = location_inventory.get(msgid, {})
            descriptions = location.get("descriptions") or []
            files = location.get("files") or []
            rows.append(
                {
                    "msgid": msgid,
                    "msgstr": entry.get("msgstr", ""),
                    "origin": entry.get("origin", "missing"),
                    "screen_hint": " | ".join(descriptions) if descriptions else "Shared catalog or backend-generated text",
                    "template_location": ", ".join(files) if files else "-",
                }
            )

        return rows

    def _translation_workbook(self, company, language="ar"):
        workbook = Workbook()
        instructions = workbook.active
        instructions.title = "Instructions"
        instructions.append(["Nezam Translation Sheet"])
        instructions.append(["Edit only the Arabic Translation column in the Translations sheet."])
        instructions.append(["Keep English Source unchanged so import can match the correct UI text."])
        instructions.append(["If Arabic Translation is left blank, any company-specific override for that row will be cleared."])

        sheet = workbook.create_sheet(self.TRANSLATION_SHEET_NAME)
        sheet.append(self.TRANSLATION_SHEET_HEADERS)
        for row in self._translation_sheet_rows(company, language):
            sheet.append(
                [
                    row["msgid"],
                    row["msgstr"],
                    row["screen_hint"],
                    row["template_location"],
                    row["origin"],
                ]
            )

        sheet.freeze_panes = "A2"
        column_widths = {
            "A": 44,
            "B": 44,
            "C": 36,
            "D": 44,
            "E": 16,
        }
        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width

        return workbook

    def download_translation_sheet(self, request, company, language="ar"):
        workbook = self._translation_workbook(company, language)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"{company.name.strip().replace(' ', '_') or 'company'}_translations_{language}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def import_translation_sheet(self, request, company, language="ar"):
        uploaded_file = request.FILES.get("translation_sheet")
        if not uploaded_file:
            messages.error(request, "Upload an Excel sheet first.")
            return self._redirect_with_scope(request, tab="translations")

        try:
            workbook = load_workbook(uploaded_file, data_only=True)
        except Exception as exc:
            messages.error(request, f"Could not read the Excel sheet: {exc}")
            return self._redirect_with_scope(request, tab="translations")

        if self.TRANSLATION_SHEET_NAME not in workbook.sheetnames:
            messages.error(request, f'The uploaded file must contain a "{self.TRANSLATION_SHEET_NAME}" sheet.')
            return self._redirect_with_scope(request, tab="translations")

        sheet = workbook[self.TRANSLATION_SHEET_NAME]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(value or "").strip() for value in (header_cells or ())]
        header_map = {header: index for index, header in enumerate(headers)}
        missing_headers = [header for header in self.TRANSLATION_SHEET_HEADERS[:2] if header not in header_map]
        if missing_headers:
            messages.error(request, "The Excel sheet is missing required columns: " + ", ".join(missing_headers))
            return self._redirect_with_scope(request, tab="translations")

        source_index = header_map["English Source"]
        target_index = header_map["Arabic Translation"]
        updated_count = 0
        cleared_count = 0
        skipped_count = 0
        row_errors = []

        existing_overrides = get_company_translation_map(company, language)
        existing_keys = {
            canonicalize_translation_key(key): key
            for key in existing_overrides.keys()
        }

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source = str(row[source_index] or "").strip()
            target = str(row[target_index] or "").strip()
            if not source:
                skipped_count += 1
                continue

            try:
                if target:
                    upsert_company_translation(company, language, source, target)
                    updated_count += 1
                else:
                    matched_key = existing_keys.get(canonicalize_translation_key(source))
                    if matched_key:
                        remove_company_translation(company, language, matched_key)
                        cleared_count += 1
                    else:
                        skipped_count += 1
            except Exception as exc:
                row_errors.append(f"row {row_number}: {exc}")

        if row_errors:
            messages.error(
                request,
                "Translation import finished with errors in "
                + ", ".join(row_errors[:5])
                + ("." if len(row_errors) <= 5 else " ..."),
            )

        messages.success(
            request,
            f"Translation sheet applied. Updated {updated_count}, cleared {cleared_count}, skipped {skipped_count}.",
        )
        return self._redirect_with_scope(request, tab="translations")

    def get(self, request):
        if not user_has_role(request.user, "ui.settings.view"):
            messages.error(request, "You are not authorized to access settings.")
            return redirect_to_role_home(request.user)

        company = require_company(request.user)
        if not company:
            return redirect('factory_setup')

        if request.GET.get("action") == "download_translation_sheet":
            return self.download_translation_sheet(request, company, "ar")

        db_alias = self._tenant_db_alias(request)
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        scope = self._get_settings_scope(request)
        team_members = list(
            User.objects.using(db_alias).filter(
            profile__company=company,
            profile__app_scope__in=self._profile_scope_values(scope)
            ).select_related('profile', 'profile__role')
        )
        self._ensure_system_roles(db_alias)
        inviter_allowed_roles = self._allowed_role_names_for_inviter(request.user)
        scope_allowed_roles = self._allowed_role_names_for_scope(scope)
        allowed_role_names = [name for name in scope_allowed_roles if name in inviter_allowed_roles]
        available_roles = list(Role.objects.using(db_alias).filter(name__in=allowed_role_names))
        role_name_by_id = {role.id: role.name for role in available_roles}

        for member in team_members:
            profile = getattr(member, 'profile', None)
            role_id = getattr(profile, 'role_id', None) if profile else None
            member.safe_role_name = role_name_by_id.get(role_id, "")
            member.department_labels = self._split_department_values(getattr(profile, "department", None) if profile else None)

        active_team_members_count = sum(1 for member in team_members if member.is_active)
        inactive_team_members_count = len(team_members) - active_team_members_count

        context = {
            'team_members': team_members,
            'team_members_count': len(team_members),
            'active_team_members_count': active_team_members_count,
            'inactive_team_members_count': inactive_team_members_count,
            'available_roles': available_roles,
            'department_options': self._department_options_for_scope(settings, scope, company, db_alias),
            'custom_department_options': self._department_catalog_for_scope(settings, scope),
            'shift_options': self._shift_options(settings),
            'shift_mode_choices': SHIFT_MODE_CHOICES,
            'translation_entries': self._translation_entries(company, "ar"),
            'translation_preview_groups': self._translation_preview_groups(company, "ar"),
            'translation_feedback': self._translation_feedback(request),
            'translation_language': 'ar',
            'translation_sheet_rows': self._translation_sheet_rows(company, "ar"),
            'company': company,
            'settings': settings,
            'normalized_shift_config': normalize_shift_configuration(
                settings.shift_configuration,
                shift_mode=getattr(settings, 'shift_mode', '3'),
                default_enabled=True,
            ),
            'active_tab': request.GET.get('tab', 'general'),
            'settings_scope': scope,
            'current_role_name': resolve_user_role(request.user) or '',
            'week_days': [
                (0, 'Monday', 'Mon'), (1, 'Tuesday', 'Tue'), (2, 'Wednesday', 'Wed'), 
                (3, 'Thursday', 'Thu'), (4, 'Friday', 'Fri'), (5, 'Saturday', 'Sat'), (6, 'Sunday', 'Sun')
            ]
        }
        return render(request, 'manufacturing/settings_dashboard_v2.html', context)

    def post(self, request):
        if not user_has_role(request.user, "ui.settings.view"):
            messages.error(request, "You are not authorized to update settings.")
            return redirect_to_role_home(request.user)

        company = require_company(request.user)
        action = request.POST.get('action')
        if action not in self.SELF_SERVICE_ACTIONS and not user_has_role(request.user, [RoleType.ADMIN, RoleType.PLANNER]):
            messages.error(request, "Only admins and planners can manage workspace settings.")
            return self._redirect_with_scope(request)

        if action == 'update_profile':
            return self.update_profile(request)
        elif action == 'update_password':
            return self.update_password(request)
        elif action == 'update_system':
            return self.update_system(request, company)
        elif action == 'add_holiday':
            return self.add_holiday(request, company)
        elif action == 'remove_holiday':
            return self.remove_holiday(request, company)
        elif action == 'update_weekly_holidays':
            return self.update_weekly_holidays(request, company)
        elif action == 'invite_user':
            return self.invite_user(request, company)
        elif action == 'update_member':
            return self.update_member(request, company)
        elif action == 'reset_member_password':
            return self.reset_member_password(request, company)
        elif action == 'delete_user':
            return self.delete_user(request, company)
        elif action == 'update_shifts':
            return self.update_shifts(request, company)
        elif action == 'update_system_rules':
            return self.update_system_rules(request, company)
        elif action == 'add_department':
            return self.add_department(request, company)
        elif action == 'remove_department':
            return self.remove_department(request, company)
        elif action == 'save_translation':
            return self.save_translation(request)
        elif action == 'delete_translation':
            return self.delete_translation(request)
        elif action == 'import_translation_sheet':
            return self.import_translation_sheet(request, company, "ar")
        
        messages.error(request, "Unknown Action")
        return self._redirect_with_scope(request)

    def update_profile(self, request):
        user = request.user
        db_alias = self._tenant_db_alias(request)
        profile, _ = Profile.objects.using(db_alias).get_or_create(user=user)
        
        if 'first_name' in request.POST:
            user.first_name = (request.POST.get('first_name') or '').strip()
        if 'last_name' in request.POST:
            user.last_name = (request.POST.get('last_name') or '').strip()
        if 'email' in request.POST:
            user.email = (request.POST.get('email') or '').strip()
        user.save(using=db_alias)
        
        if 'phone' in request.POST:
            profile.phone = (request.POST.get('phone') or '').strip()
        
        if 'profile_image' in request.FILES:
            profile.profile_image = request.FILES['profile_image']
            
        profile.save(using=db_alias)

        company = require_company(user)
        if company:
            if 'company_logo' in request.FILES:
                company.logo = request.FILES['company_logo']
            if 'company_name' in request.POST:
                name = (request.POST.get('company_name') or '').strip()
                if name:
                    company.name = name
            if 'email' in request.POST:
                company.support_email = (request.POST.get('email') or '').strip()
            company.save(using=db_alias)
        
        messages.success(request, "Profile Updated Successfully")
        return self._redirect_with_scope(request)
    
    def update_password(self, request):
        user = request.user
        db_alias = self._tenant_db_alias(request)
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not user.check_password(current_password):
            messages.error(request, "Incorrect current password.")
            return self._redirect_with_scope(request)
        
        if new_password != confirm_password:
            messages.error(request, "New passwords do not match.")
            return self._redirect_with_scope(request)

        try:
            validate_password(new_password, user=user)
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
            return self._redirect_with_scope(request)
            
        user.set_password(new_password)
        user.save(using=db_alias)
        update_session_auth_hash(request, user) # Keep user logged in
        messages.success(request, "Password updated successfully.")
        return self._redirect_with_scope(request)

    def update_system(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        settings.theme = request.POST.get('theme', 'light')
        settings.language = request.POST.get('language', 'en')
        settings.save()
        messages.success(request, "System settings updated.")
        return self._redirect_with_scope(request, tab="system")

    def add_holiday(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        date = request.POST.get('date')
        name = request.POST.get('name')
        
        if date and name:
            holidays = settings.holidays or []
            holidays.append({'date': date, 'name': name})
            settings.holidays = holidays
            settings.save()
            messages.success(request, "Holiday added.")
        return self._redirect_with_scope(request, tab="shift-planner")
        
    def remove_holiday(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        holiday_idx = request.POST.get('holiday_idx')
        
        if holiday_idx is not None:
            try:
                idx = int(holiday_idx)
                holidays = settings.holidays or []
                if 0 <= idx < len(holidays):
                    holidays.pop(idx)
                    settings.holidays = holidays
                    settings.save()
                    messages.success(request, "Holiday removed.")
            except:
                messages.error(request, "Invalid holiday index.")
        
        return self._redirect_with_scope(request, tab="shift-planner")

    def update_weekly_holidays(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        # Expecting a list of day indices from checkboxes
        selected_days = request.POST.getlist('weekly_holidays')
        settings.weekly_holidays = [int(d) for d in selected_days]
        settings.save()
        messages.success(request, "Weekly holidays updated.")
        return self._redirect_with_scope(request, tab="shift-planner")

    def update_shifts(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        shift_mode = normalize_shift_mode(request.POST.get('shift_mode'), default=getattr(settings, 'shift_mode', '3'))
        enabled_shift_keys = enabled_shift_keys_for_mode(shift_mode)
        morning_start = request.POST.get('morning_start') or '06:00'
        morning_end = request.POST.get('morning_end') or '14:00'
        evening_start = request.POST.get('evening_start') or '14:00'
        evening_end = request.POST.get('evening_end') or '22:00'
        night_start = request.POST.get('night_start') or '22:00'
        night_end = request.POST.get('night_end') or '06:00'
        
        shift_config = {
            'morning': {'start': morning_start, 'end': morning_end, 'enabled': 'morning' in enabled_shift_keys},
            'evening': {'start': evening_start, 'end': evening_end, 'enabled': 'afternoon' in enabled_shift_keys},
            'night': {'start': night_start, 'end': night_end, 'enabled': 'night' in enabled_shift_keys}
        }
        
        settings.shift_mode = shift_mode
        settings.shift_configuration = shift_config
        settings.save()
        messages.success(request, "Shift schedule updated.")
        return self._redirect_with_scope(request, tab="shift-planner")

    def update_system_rules(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        
        # Checkboxes only send value if checked
        settings.auto_assign_workers = request.POST.get('auto_assign_workers') == 'on'
        settings.predictive_maintenance = request.POST.get('predictive_maintenance') == 'on'
        settings.auto_fault_lockdown = request.POST.get('auto_fault_lockdown') == 'on'
        settings.trouble_ticket_integration = request.POST.get('trouble_ticket_integration') == 'on'
        operation_flow_mode = (request.POST.get('default_operation_flow_mode') or 'series').strip().lower()
        if operation_flow_mode not in {'series', 'parallel'}:
            operation_flow_mode = 'series'
        settings.default_operation_flow_mode = operation_flow_mode
        
        settings.save()
        messages.success(request, "System configuration updated.")
        return self._redirect_with_scope(request, tab="system")

    def add_department(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        scope = self._get_settings_scope(request)
        department_name = (request.POST.get("department_name") or "").strip()

        if not department_name:
            messages.error(request, "Department name is required.")
            return self._redirect_with_scope(request, tab="members")

        catalog = settings.department_catalog or {}
        scope_departments = catalog.get(scope, [])
        if scope == "manufacturing" and not scope_departments:
            scope_departments = catalog.get("planner", [])
        if not isinstance(scope_departments, list):
            scope_departments = []

        if any(existing.lower() == department_name.lower() for existing in scope_departments):
            messages.error(request, f"{department_name} already exists.")
            return self._redirect_with_scope(request, tab="members")

        scope_departments.append(department_name)
        catalog[scope] = scope_departments
        if scope == "manufacturing":
            catalog.pop("planner", None)
        settings.department_catalog = catalog
        settings.save(update_fields=["department_catalog"])

        messages.success(request, f"Department {department_name} added.")
        return self._redirect_with_scope(request, tab="members")

    def remove_department(self, request, company):
        settings, _ = SystemSettings.objects.get_or_create(company=company)
        scope = self._get_settings_scope(request)
        department_name = (request.POST.get("department_name") or "").strip()

        if not department_name:
            messages.error(request, "Department name is required.")
            return self._redirect_with_scope(request, tab="members")

        catalog = settings.department_catalog or {}
        scope_departments = catalog.get(scope, [])
        if scope == "manufacturing" and not scope_departments:
            scope_departments = catalog.get("planner", [])
        if not isinstance(scope_departments, list):
            scope_departments = []

        filtered = [item for item in scope_departments if str(item or "").strip().lower() != department_name.lower()]
        if len(filtered) == len(scope_departments):
            messages.error(request, f"{department_name} was not found in custom departments.")
            return self._redirect_with_scope(request, tab="members")

        catalog[scope] = filtered
        if scope == "manufacturing":
            catalog.pop("planner", None)
        settings.department_catalog = catalog
        settings.save(update_fields=["department_catalog"])

        messages.success(request, f"Department {department_name} removed.")
        return self._redirect_with_scope(request, tab="members")

    def invite_user(self, request, company):
        db_alias = self._tenant_db_alias(request)
        self._ensure_system_roles(db_alias)
        scope = self._get_settings_scope(request)
        first_name = (request.POST.get('first_name') or '').strip()
        last_name = (request.POST.get('last_name') or '').strip()
        email = (request.POST.get('email') or '').strip().lower()
        phone = self._normalize_member_phone(request.POST.get('phone'))
        department_values = self._split_department_values(request.POST.getlist('department'))
        shift = (request.POST.get('shift') or '').strip().lower()
        worker_mode_enabled = request.POST.get('worker_mode_enabled') == 'on'
        password = request.POST.get('password') or ''
        confirm_password = request.POST.get('confirm_password') or ''
        role_id = request.POST.get('role')

        settings, _ = SystemSettings.objects.get_or_create(company=company)
        department_options = self._department_options_for_scope(settings, scope, company, db_alias)

        if not all([first_name, last_name]) or not (email or phone):
            error_message = "Complete first name, last name, and at least one contact method."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "1", "invite_error": error_message},
            )

        if not role_id:
            error_message = "Choose a role before continuing."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "2", "invite_error": error_message},
            )

        role = get_object_or_404(Role.objects.using(db_alias), pk=role_id)
        is_planner_role = role.name == RoleType.PLANNER.value
        skips_department = self._role_skips_department(role.name)
        if is_planner_role:
            department_values = list(department_options)
        elif role.name == RoleType.STORE.value:
            department_values = []
        elif not department_options:
            error_message = "Create at least one department/category first, then assign it to this team member."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "2", "invite_error": error_message},
            )
        elif not department_values:
            error_message = "Select at least one coverage category before continuing."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "2", "invite_error": error_message},
            )

        if not all([password, confirm_password]):
            error_message = "Enter and confirm the primary password to finish account creation."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "3", "invite_error": error_message},
            )

        if password != confirm_password:
            error_message = "Password and confirm password do not match."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "3", "invite_error": error_message},
            )

        if len(password) < 8:
            error_message = "Password must be at least 8 characters."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "3", "invite_error": error_message},
            )

        try:
            validate_password(password)
        except ValidationError as exc:
            error_message = " ".join(exc.messages)
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "3", "invite_error": error_message},
            )

        if email and User.objects.using(db_alias).filter(email__iexact=email).exists():
            error_message = f"User with email {email} already exists."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "1", "invite_error": error_message},
            )

        if phone and Profile.objects.using(db_alias).filter(phone=phone).exists():
            error_message = f"User with phone {phone} already exists."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "1", "invite_error": error_message},
            )

        username = self._build_unique_member_username(db_alias, company, email=email, phone=phone)
        new_user = User.objects.db_manager(db_alias).create_user(username=username, email=email, password=password)
        new_user.first_name = first_name
        new_user.last_name = last_name
        new_user.save(using=db_alias)

        inviter_allowed_roles = self._allowed_role_names_for_inviter(request.user)
        scope_allowed_roles = self._allowed_role_names_for_scope(scope)
        allowed_role_names = [name for name in scope_allowed_roles if name in inviter_allowed_roles]
        if role.name not in allowed_role_names:
            new_user.delete(using=db_alias)
            error_message = "You are not allowed to invite users with this role."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "2", "invite_error": error_message},
            )

        valid_department_keys = {
            option.lower() for option in department_options
        }
        invalid_departments = [] if skips_department else [value for value in department_values if value.lower() not in valid_department_keys]
        if invalid_departments:
            error_message = "Invalid department/category selection."
            messages.error(request, error_message)
            return self._redirect_with_scope(
                request,
                tab="members",
                extra_params={"member_modal": "invite", "member_step": "2", "invite_error": error_message},
            )

        # Signal already created the profile, so we just get and update it
        profile, created = Profile.objects.using(db_alias).get_or_create(user=new_user)
        profile.company = company
        profile.role = role
        profile.app_scope = scope
        profile.department = self._join_department_values(department_values) if department_values else None
        profile.shift = shift or None
        profile.phone = phone or None
        profile.worker_mode_enabled = worker_mode_enabled if role.name == RoleType.SUPERVISOR.value else False
        profile.save(using=db_alias)
        
        messages.success(request, f"User {first_name} created successfully.")
        return self._redirect_with_scope(request, tab="members")

    def update_member(self, request, company):
        db_alias = self._tenant_db_alias(request)
        scope = self._get_settings_scope(request)
        member_id = request.POST.get("member_id")
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        phone = self._normalize_member_phone(request.POST.get("phone"))
        department_values = self._split_department_values(request.POST.getlist("department"))
        shift = (request.POST.get("shift") or "").strip().lower()
        worker_mode_enabled = request.POST.get("worker_mode_enabled") == "on"
        role_id = request.POST.get("role")

        if not member_id or not all([first_name, last_name, role_id]) or not (email or phone):
            messages.error(request, "First name, last name, role, and at least one contact method are required.")
            return self._redirect_with_scope(request, tab="members")

        member = get_object_or_404(
            User.objects.using(db_alias),
            pk=member_id,
            profile__company=company,
            profile__app_scope__in=self._profile_scope_values(scope),
        )

        if member == request.user:
            messages.error(request, "Edit your own account from the profile section.")
            return self._redirect_with_scope(request, tab="members")

        if email and User.objects.using(db_alias).filter(email__iexact=email).exclude(pk=member.pk).exists():
            messages.error(request, f"User with email {email} already exists.")
            return self._redirect_with_scope(request, tab="members")

        role = get_object_or_404(Role.objects.using(db_alias), pk=role_id)
        inviter_allowed_roles = self._allowed_role_names_for_inviter(request.user)
        scope_allowed_roles = self._allowed_role_names_for_scope(scope)
        allowed_role_names = [name for name in scope_allowed_roles if name in inviter_allowed_roles]
        if role.name not in allowed_role_names:
            messages.error(request, "You are not allowed to assign this role.")
            return self._redirect_with_scope(request, tab="members")

        settings, _ = SystemSettings.objects.get_or_create(company=company)
        department_options = self._department_options_for_scope(settings, scope, company, db_alias)
        is_planner_role = role.name == RoleType.PLANNER.value
        skips_department = self._role_skips_department(role.name)
        if is_planner_role:
            department_values = list(department_options)
        elif role.name == RoleType.STORE.value:
            department_values = []
        elif not department_options:
            messages.error(request, "Create at least one department/category first, then assign it to this team member.")
            return self._redirect_with_scope(request, tab="members")
        elif not department_values:
            messages.error(request, "Select at least one coverage category.")
            return self._redirect_with_scope(request, tab="members")

        valid_department_keys = {
            option.lower() for option in department_options
        }
        invalid_departments = [] if skips_department else [value for value in department_values if value.lower() not in valid_department_keys]
        if invalid_departments:
            messages.error(request, "Invalid department/category selection.")
            return self._redirect_with_scope(request, tab="members")

        valid_shifts = {item["value"] for item in self._shift_options(settings)}
        if shift and shift not in valid_shifts:
            messages.error(request, "Invalid shift selection.")
            return self._redirect_with_scope(request, tab="members")

        profile, _created = Profile.objects.using(db_alias).get_or_create(user=member)
        if phone and Profile.objects.using(db_alias).filter(phone=phone).exclude(user_id=member.pk).exists():
            messages.error(request, f"User with phone {phone} already exists.")
            return self._redirect_with_scope(request, tab="members")

        member.first_name = first_name
        member.last_name = last_name
        member.email = email
        member.save(using=db_alias)

        profile.company = company
        profile.role = role
        profile.app_scope = scope
        profile.department = self._join_department_values(department_values) if department_values else None
        profile.shift = shift or None
        profile.phone = phone or None
        profile.worker_mode_enabled = worker_mode_enabled if role.name == RoleType.SUPERVISOR.value else False
        profile.save(using=db_alias)

        messages.success(request, f"{first_name} updated successfully.")
        return self._redirect_with_scope(request, tab="members")

    def reset_member_password(self, request, company):
        if not self._can_admin_reset_member_password(request.user):
            messages.error(request, "Only admins can reset other users' passwords.")
            return self._redirect_with_scope(request, tab="members")

        db_alias = self._tenant_db_alias(request)
        scope = self._get_settings_scope(request)
        member_id = request.POST.get("member_id")
        new_password = request.POST.get("new_password") or ""
        confirm_password = request.POST.get("confirm_password") or ""

        if not member_id or not new_password or not confirm_password:
            messages.error(request, "Enter and confirm the new password.")
            return self._redirect_with_scope(request, tab="members")

        member = get_object_or_404(
            User.objects.using(db_alias),
            pk=member_id,
            profile__company=company,
            profile__app_scope__in=self._profile_scope_values(scope),
        )

        if member == request.user:
            messages.error(request, "Use the profile password form to change your own password.")
            return self._redirect_with_scope(request, tab="members")

        if new_password != confirm_password:
            messages.error(request, "New passwords do not match.")
            return self._redirect_with_scope(request, tab="members")

        if len(new_password) < 8:
            messages.error(request, "Password must be at least 8 characters.")
            return self._redirect_with_scope(request, tab="members")

        try:
            validate_password(new_password, user=member)
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
            return self._redirect_with_scope(request, tab="members")

        member.set_password(new_password)
        member.save(using=db_alias)
        messages.success(request, f"Password reset for {member.get_full_name() or member.username}.")
        return self._redirect_with_scope(request, tab="members")

    def delete_user(self, request, company):
        user_id = request.POST.get('user_id')
        scope = self._get_settings_scope(request)
        db_alias = self._tenant_db_alias(request)
        user_to_delete = get_object_or_404(
            User.objects.using(db_alias),
            pk=user_id,
            profile__company=company,
            profile__app_scope__in=self._profile_scope_values(scope),
        )
        
        if user_to_delete == request.user:
            messages.error(request, "You cannot delete yourself.")
            return self._redirect_with_scope(request, tab="members")

        target_profile = Profile.objects.using(db_alias).filter(user_id=user_to_delete.id).select_related("role").first()
        target_role = getattr(getattr(target_profile, "role", None), "name", "")
        if target_role not in self._allowed_role_names_for_inviter(request.user):
            messages.error(request, "You are not allowed to delete users with this role.")
            return self._redirect_with_scope(request, tab="members")
            
        user_to_delete.delete(using=db_alias)
        messages.success(request, "User removed from company.")
        return self._redirect_with_scope(request, tab="members")

    def save_translation(self, request):
        company = require_company(request.user)
        language = (request.POST.get("language") or "ar").strip().lower()
        msgid = (request.POST.get("msgid") or "").strip()
        msgstr = (request.POST.get("msgstr") or "").strip()

        if language != "ar":
            messages.error(request, "Only Arabic translation editing is enabled right now.")
            return self._redirect_with_scope(request, tab="translations")

        if not msgid:
            messages.error(request, "Source text is required.")
            return self._redirect_with_scope(request, tab="translations")

        if not msgstr:
            messages.error(request, "Translated text is required.")
            return self._redirect_with_scope(request, tab="translations")

        try:
            if not company:
                raise ValueError("Company context is required.")
            upsert_company_translation(company, language, msgid, msgstr)
        except Exception as exc:
            messages.error(request, f"Failed to save translation: {exc}")
            return self._redirect_with_scope(request, tab="translations")

        runtime_value = resolve_runtime_translation(company, msgid, language)
        known_sources = self._known_translation_sources(company, language)
        exact_known = msgid in known_sources
        canonical_known = any(
            canonicalize_translation_key(source) == canonicalize_translation_key(msgid)
            for source in known_sources
        )
        closest_matches = get_close_matches(msgid, known_sources, n=4, cutoff=0.55)
        request.session["translation_feedback"] = {
            "msgid": msgid,
            "msgstr": msgstr,
            "runtime_value": runtime_value,
            "exact_known": exact_known,
            "canonical_known": canonical_known,
            "closest_matches": closest_matches,
            "applied": runtime_value == msgstr,
        }

        applied = runtime_value == msgstr
        request.session["translation_feedback"]["applied"] = applied

        messages.success(request, f"Translation saved for: {msgid}")
        if not exact_known and not applied:
            if closest_matches:
                messages.warning(
                    request,
                    "Saved, but no exact UI field uses this source yet. Try one of: "
                    + ", ".join(closest_matches[:3]),
                )
            else:
                messages.warning(
                    request,
                    "Saved, but no exact UI field uses this source yet. Use the preset runtime keys below.",
                )
        return self._redirect_with_scope(request, tab="translations")

    def delete_translation(self, request):
        company = require_company(request.user)
        language = (request.POST.get("language") or "ar").strip().lower()
        msgid = (request.POST.get("msgid") or "").strip()

        if language != "ar":
            messages.error(request, "Only Arabic translation editing is enabled right now.")
            return self._redirect_with_scope(request, tab="translations")

        if not msgid:
            messages.error(request, "Source text is required.")
            return self._redirect_with_scope(request, tab="translations")

        try:
            if not company:
                raise ValueError("Company context is required.")
            remove_company_translation(company, language, msgid)
        except Exception as exc:
            messages.error(request, f"Failed to delete translation: {exc}")
            return self._redirect_with_scope(request, tab="translations")

        messages.success(request, f"Translation removed: {msgid}")
        return self._redirect_with_scope(request, tab="translations")
