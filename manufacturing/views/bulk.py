import os
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views import View

from accounts.constants import RoleType
from accounts.models import Profile, Role
from manufacturing.forms import BulkImportForm
from manufacturing.import_catalog import get_bulk_import_filenames
from manufacturing.models import (
    BillOfMaterial,
    BOMComponent,
    BOMOperation,
    BOMOperationMaterial,
    Company,
    Machine,
    Product,
    ProductionStage,
    SystemSettings,
    WorkOrder,
)
from tenancy.context import reset_current_tenant_db, set_current_tenant_db
from tenancy.db import ensure_tenant_database_ready
from tenancy.models import Tenant
from .dashboard import require_company, user_has_role
from manufacturing.services import (
    NotificationService,
    flag_bom_change_impact,
    get_company_default_operation_flow_mode,
)
from manufacturing.utils import normalize_machine_code


class BulkImportView(LoginRequiredMixin, View):
    def get(self, request):
        if not user_has_role(request.user, "ui.bulk_import.manage"):
            return redirect('planner_dashboard')
        return render(request, 'manufacturing/bulk_import.html', {'form': BulkImportForm()})


class DownloadTemplateView(LoginRequiredMixin, View):
    def get(self, request, filename):
        allowed = get_bulk_import_filenames()
        if filename not in allowed:
            raise Http404("Template not found")

        file_path = os.path.join(settings.BASE_DIR, 'templates', 'downloads', filename)
        if not os.path.exists(file_path):
            raise Http404("File not on server")

        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response


class HandleBulkImportView(LoginRequiredMixin, View):
    ONBOARDING_NEXT_TARGETS = {"onboarding_data", "onboarding_users", "dashboard"}

    @staticmethod
    def _notify_import_result(request, singular_label, plural_label, count, errors, empty_hint):
        if errors:
            messages.warning(
                request,
                f"Imported {count} {plural_label}. {len(errors)} errors occurred."
            )
            for err in errors[:5]:
                messages.warning(request, err)
            return

        if count <= 0:
            messages.warning(request, f"No {plural_label} were imported. {empty_hint}")
            return

        label = singular_label if count == 1 else plural_label
        messages.success(request, f"Successfully imported {count} {label}.")

    @staticmethod
    def _redirect_to_next(request, fallback_name):
        next_target = (request.POST.get("next") or request.GET.get("next") or fallback_name or "").strip()
        if not next_target:
            return redirect(fallback_name)
        return redirect(next_target)

    def _mark_onboarding_planner_reset_if_needed(self, request):
        next_target = (request.POST.get("next") or request.GET.get("next") or "").strip()
        if next_target in self.ONBOARDING_NEXT_TARGETS:
            request.session["reset_planner_workspace_state"] = True

    def _resolve_tenant_alias(self, request):
        alias = getattr(request, "tenant_db_alias", None)
        if alias and alias != "default":
            return alias

        tenant_code = (
            request.POST.get("tenant_code")
            or request.GET.get("tenant")
            or request.session.get("tenant_code")
        )
        if not tenant_code:
            return None

        tenant = Tenant.objects.using("default").filter(code=tenant_code, is_active=True).first()
        if not tenant:
            return None
        return ensure_tenant_database_ready(tenant)

    def _resolve_company(self, request, db_alias):
        profile = Profile.objects.using(db_alias).filter(user_id=request.user.id).first()
        if profile and profile.company_id:
            company = Company.objects.using(db_alias).filter(id=profile.company_id).first()
            if company:
                return company

        company = Company.objects.using(db_alias).order_by("-created_at").first()
        if not company:
            return None

        profile, _ = Profile.objects.using(db_alias).get_or_create(user_id=request.user.id)
        if profile.company_id != company.id:
            profile.company_id = company.id
            profile.save(using=db_alias, update_fields=["company"])
        return company

    def post(self, request):
        db_alias = self._resolve_tenant_alias(request)
        if not db_alias:
            messages.error(request, "Session expired. Please sign in again using your company code.")
            return redirect("login")

        ctx_token = set_current_tenant_db(db_alias)
        try:
            company = self._resolve_company(request, db_alias)
            if not company:
                messages.error(request, "No company assigned to your account. Please contact administrator.")
                return self._redirect_to_next(request, 'bulk_import_dashboard')
            if not user_has_role(request.user, "ui.bulk_import.manage"):
                messages.error(request, "You are not authorized to bulk import production data.")
                return self._redirect_to_next(request, 'planner_dashboard')

            form = BulkImportForm(request.POST, request.FILES)
            if not form.is_valid():
                messages.error(request, "Invalid form submission.")
                return self._redirect_to_next(request, 'bulk_import_dashboard')

            import_type = form.cleaned_data['import_type']
            file = request.FILES['file']
            user_manager = User.objects.db_manager(db_alias)
            role_manager = Role.objects.using(db_alias)
            profile_manager = Profile.objects.using(db_alias)
            company_manager = Company.objects.using(db_alias)
            machine_manager = Machine.objects.using(db_alias)
            product_manager = Product.objects.using(db_alias)
            stage_manager = ProductionStage.objects.using(db_alias)
            bom_manager = BillOfMaterial.objects.using(db_alias)
            bom_component_manager = BOMComponent.objects.using(db_alias)
            bom_operation_manager = BOMOperation.objects.using(db_alias)
            bom_operation_material_manager = BOMOperationMaterial.objects.using(db_alias)
            work_order_manager = WorkOrder.objects.using(db_alias)

            def _coerce_datetime(value):
                if value is None or value == '':
                    return None
                if isinstance(value, datetime):
                    dt = value
                elif isinstance(value, date):
                    dt = datetime.combine(value, time.min)
                elif isinstance(value, str):
                    dt = parse_datetime(value)
                    if not dt:
                        d = parse_date(value)
                        dt = datetime.combine(d, time.min) if d else None
                else:
                    dt = None
                if dt and timezone.is_naive(dt):
                    dt = timezone.make_aware(dt, timezone.get_current_timezone())
                return dt

            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                count = 0
                errors = []
                header_values = [str(cell.value or "").strip() for cell in sheet[1]]

                def _normalize_header(value):
                    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())

                normalized_headers = {
                    _normalize_header(value): index
                    for index, value in enumerate(header_values)
                    if str(value or "").strip()
                }

                def _row_value(row, candidate_headers, fallback_index=None):
                    for header in candidate_headers:
                        key = _normalize_header(header)
                        if key in normalized_headers and normalized_headers[key] < len(row):
                            return row[normalized_headers[key]]
                    for existing_key, index in normalized_headers.items():
                        if index >= len(row):
                            continue
                        if any(_normalize_header(header) in existing_key for header in candidate_headers):
                            return row[index]
                    if not normalized_headers and fallback_index is not None and fallback_index < len(row):
                        return row[fallback_index]
                    return None

                def _clean_text(value):
                    return " ".join(str(value or "").strip().split())

                def _is_blank(value):
                    return value is None or _clean_text(value) == ""

                def _parse_decimal(value):
                    if _is_blank(value):
                        return None
                    try:
                        return Decimal(str(value).strip())
                    except (InvalidOperation, ValueError, TypeError):
                        return None

                def _coerce_bool(value):
                    if isinstance(value, bool):
                        return value
                    raw = str(value or "").strip().lower()
                    return raw in {"1", "true", "yes", "y", "on", "enabled"}

                def _coerce_date(value):
                    if value is None or value == "":
                        return None
                    if isinstance(value, datetime):
                        return value.date()
                    if isinstance(value, date):
                        return value
                    if isinstance(value, str):
                        return parse_date(value.strip())
                    return None

                def _coerce_decimal(value, default="0"):
                    if value is None or value == "":
                        value = default
                    try:
                        return Decimal(str(value).strip())
                    except (InvalidOperation, ValueError, TypeError):
                        return Decimal(str(default))

                def _coerce_positive_int(value, default=1):
                    try:
                        return max(int(_coerce_decimal(value, default)), 1)
                    except Exception:
                        return max(int(default or 1), 1)

                def _normalize_status(value, allowed, default):
                    status = _clean_text(value).lower()
                    return status if status in allowed else default

                def _has_any_header(candidate_headers):
                    return any(_normalize_header(header) in normalized_headers for header in candidate_headers)

                def _matching_headers(candidate_headers):
                    candidate_keys = {_normalize_header(header) for header in candidate_headers}
                    return {
                        header
                        for header in normalized_headers
                        if header in candidate_keys
                    }

                def _has_header_group(candidate_headers):
                    return bool(_matching_headers(candidate_headers))

                def _reject_if_missing_template_signature(label, required_header_groups, hint):
                    if all(_has_header_group(group) for group in required_header_groups):
                        return False
                    messages.error(request, f"The selected file does not look like a {label} template. {hint}")
                    return True

                def _first_present(row, candidate_headers, fallback_index=None, carry_value=None):
                    value = _row_value(row, candidate_headers, fallback_index)
                    if value is None or _clean_text(value) == "":
                        return carry_value
                    return value

                def _normalize_unit(value, default="pcs"):
                    raw = _clean_text(value).lower()
                    aliases = {
                        "piece": "pcs",
                        "pieces": "pcs",
                        "pc": "pcs",
                        "pcs": "pcs",
                        "unit": "pcs",
                        "units": "pcs",
                        "kg": "kg",
                        "kgs": "kg",
                        "kilogram": "kg",
                        "kilograms": "kg",
                        "g": "gm",
                        "gm": "gm",
                        "gram": "gm",
                        "grams": "gm",
                        "m": "m",
                        "meter": "m",
                        "meters": "m",
                        "cm": "cm",
                        "centimeter": "cm",
                        "centimeters": "cm",
                        "l": "l",
                        "ltr": "l",
                        "liter": "l",
                        "litre": "l",
                    }
                    return aliases.get(raw, default)

                def _normalize_scrap_type(value):
                    raw = _clean_text(value).lower().replace("-", " ").replace("_", " ")
                    if raw in {"irretrievable", "lost", "loss", "waste", "consumed"}:
                        return "irretrievable"
                    if raw in {"return", "return stock", "return to stock", "reuse", "re use", "reusable"}:
                        return "return_to_stock"
                    if raw in {"sell", "scrap", "sell scrap", "sell as scrap", "recover", "recovery"}:
                        return "sell_as_scrap"
                    return "sell_as_scrap"

                def _normalize_app_scope(value):
                    raw = _clean_text(value).lower()
                    aliases = {
                        "manufacturing": "manufacturing",
                        "production": "manufacturing",
                        "planner": "manufacturing",
                        "quality": "quality",
                        "maintenance": "maintenance",
                        "store": "store",
                    }
                    return aliases.get(raw, "")

                def _resolve_machine(raw_machine_value):
                    machine_ref = _clean_text(raw_machine_value)
                    if not machine_ref:
                        return None

                    normalized_ref = normalize_machine_code(machine_ref)
                    machine = machine_manager.filter(company=company, code__iexact=normalized_ref).first()
                    if machine:
                        return machine

                    machine = machine_manager.filter(company=company, code__iexact=machine_ref).first()
                    if machine:
                        return machine

                    return machine_manager.filter(company=company, name__iexact=machine_ref).first()

                def _next_bom_version(product):
                    max_major = 1
                    max_minor = -1
                    version_re = re.compile(r"^v?(\d+)(?:\.(\d+))?$", re.IGNORECASE)
                    for version in bom_manager.filter(product=product).values_list("version", flat=True):
                        match = version_re.match(str(version or "").strip())
                        if not match:
                            continue
                        major = int(match.group(1))
                        minor = int(match.group(2) or 0)
                        if (major, minor) > (max_major, max_minor):
                            max_major, max_minor = major, minor
                    if max_minor < 0:
                        return "v1.0"
                    max_minor += 1
                    if max_minor >= 10:
                        max_major += 1
                        max_minor = 0
                    return f"v{max_major}.{max_minor}"

                if import_type == 'products':
                    product_name_headers = ["Name", "Product Name"]
                    product_specific_headers = [
                        "Material Type",
                        "Material Type (raw/finished)",
                        "Unit",
                        "Unit (pcs/kg/m)",
                        "UOM",
                        "Description",
                        "Cost",
                    ]
                    employee_signature_headers = [
                        "Employee Name",
                        "Email",
                        "Phone",
                        "Role",
                        "Department",
                        "Departments",
                        "Department(s)",
                        "App Scope",
                        "Worker Mode Enabled",
                        "Planned Shift",
                    ]
                    has_product_name = _has_any_header(["Product Name"])
                    has_generic_name = _has_any_header(["Name"])
                    has_product_specific_header = _has_any_header(product_specific_headers)
                    has_employee_signature = bool(_matching_headers(employee_signature_headers))

                    if has_employee_signature:
                        messages.error(
                            request,
                            "This looks like a Team/Employees file, not a Products file. Select Employees or upload the Products template.",
                        )
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    if _reject_if_missing_template_signature(
                        "Products",
                        [
                            product_name_headers,
                            product_specific_headers,
                        ],
                        "Use the Products template and keep Name plus product fields like Material Type or Unit.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    if not (has_product_name or has_generic_name) or (has_generic_name and not has_product_name and not has_product_specific_header):
                        messages.error(
                            request,
                            "The selected file does not look like a Products template. Use the Products template and keep Name plus product fields like Material Type or Unit.",
                        )
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        product_name = _clean_text(_row_value(row, ["Name", "Product Name"], 0))
                        if not product_name:
                            continue
                        try:
                            material_type = _clean_text(
                                _row_value(row, ["Material Type", "Material Type (raw/finished)", "Type"], 1)
                            ).lower() or "raw"
                            if material_type not in {'raw', 'semi', 'finished', 'packaging'}:
                                material_type = 'raw'

                            unit = _clean_text(_row_value(row, ["Unit", "Unit (pcs/kg/m)", "UOM"], 2)) or 'pcs'
                            description = _clean_text(_row_value(row, ["Description"], 3))

                            existing = product_manager.filter(company=company, name__iexact=product_name).first()
                            if existing:
                                existing.name = product_name
                                existing.material_type = material_type
                                existing.unit = unit
                                existing.description = description
                                existing.save(using=db_alias, update_fields=['name', 'material_type', 'unit', 'description'])
                            else:
                                product_manager.create(
                                    company=company,
                                    name=product_name,
                                    material_type=material_type,
                                    unit=unit,
                                    description=description
                                )
                            count += 1
                        except Exception as exc:
                            errors.append(f"Row {row_idx}: Error importing '{product_name}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="product",
                        plural_label="products",
                        count=count,
                        errors=errors,
                        empty_hint="Add at least one filled product row to the Excel sheet before uploading.",
                    )
                    self._mark_onboarding_planner_reset_if_needed(request)

                elif import_type == 'machines':
                    if _reject_if_missing_template_signature(
                        "Machines",
                        [
                            ["Name", "Machine Name"],
                            ["Code"],
                            ["Status", "Type"],
                        ],
                        "Use the Machines template and keep Name, Code, Status, and Type columns.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row[0]:
                            continue
                        name = str(row[0]).strip()
                        code = str(row[1]).strip() if row[1] is not None else f"M{row_idx:03d}"
                        if not code:
                            code = f"M{row_idx:03d}"
                        status = (row[2] or 'operational')
                        if isinstance(status, str):
                            status = status.strip().lower()
                        status_aliases = {
                            'active': 'operational',
                            'running': 'operational',
                            'down': 'broken',
                        }
                        status = status_aliases.get(status, status)
                        if status not in {'operational', 'maintenance', 'broken', 'inactive'}:
                            status = 'operational'
                        is_active = status != 'inactive'
                        m_type = str(row[3]).strip() if row[3] is not None else 'General'
                        if not m_type:
                            m_type = 'General'
                        try:
                            machine_manager.update_or_create(
                                company=company,
                                code=code,
                                defaults={
                                    'name': name,
                                    'status': status,
                                    'type': m_type,
                                    'category': m_type,
                                    'is_active': is_active,
                                }
                            )
                            count += 1
                        except Exception as exc:
                            errors.append(f"Row {row_idx}: Error importing machine '{name}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="machine",
                        plural_label="machines",
                        count=count,
                        errors=errors,
                        empty_hint="Add at least one filled machine row to the Excel sheet before uploading.",
                    )
                    self._mark_onboarding_planner_reset_if_needed(request)

                elif import_type == 'employees':
                    if _reject_if_missing_template_signature(
                        "Employees",
                        [
                            ["Employee Name", "Name", "Full Name"],
                            ["Email", "Work Email"],
                            ["Role", "Role (Planner/Supervisor/Worker)"],
                        ],
                        "Use the Employees template and keep Employee Name, Email, and Role columns.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    new_departments = {}

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        name = _row_value(row, ["Employee Name", "Name", "Full Name"], 0)
                        email = _clean_text(_row_value(row, ["Email", "Work Email"], 1)).lower()
                        role_name = _row_value(row, ["Role", "Role (Planner/Supervisor/Worker)"], 2)
                        phone = _clean_text(_row_value(row, ["Phone", "Mobile", "Phone Number"]))
                        department = _clean_text(_row_value(row, ["Department", "Departments", "Department(s)", "Coverage"], 4))
                        shift = _clean_text(_row_value(row, ["Shift", "Default Shift", "Current Shift"], 5)).lower()
                        app_scope = _normalize_app_scope(_row_value(row, ["App Scope", "Scope"], 6)) or "manufacturing"
                        worker_mode_enabled = _coerce_bool(_row_value(row, ["Worker Mode", "Worker Mode Enabled"], 7))
                        planned_shift = _clean_text(_row_value(row, ["Planned Shift", "Future Shift"], 8)).lower()
                        planned_shift_start_date = _coerce_date(
                            _row_value(row, ["Planned Shift Start Date", "Shift Start Date", "Effective Start Date"], 9)
                        )
                        password = str(_row_value(row, ["Password", "Initial Password"], 10) or "Password123!").strip() or "Password123!"

                        name = _clean_text(name)
                        username = email or phone
                        if not username or not name:
                            continue
                        if user_manager.filter(username=username).exists():
                            continue

                        role_raw = (role_name or 'worker')
                        role_key = role_raw.strip().lower() if isinstance(role_raw, str) else 'worker'
                        if role_key not in {r.value for r in RoleType}:
                            role_key = 'worker'
                        if app_scope not in {"manufacturing", "quality", "maintenance", "store"}:
                            app_scope = "manufacturing"
                        if role_key == RoleType.STORE.value:
                            app_scope = "store"
                            department = ""
                        elif role_key == RoleType.PLANNER.value:
                            department = ""
                        if shift and shift not in {"morning", "evening", "night"}:
                            shift = ""
                        if planned_shift and planned_shift not in {"morning", "evening", "night"}:
                            planned_shift = ""
                            planned_shift_start_date = None

                        try:
                            first_name, _, last_name = name.partition(" ")
                            user = user_manager.create_user(
                                username=username,
                                email=email,
                                password=password,
                                first_name=first_name or name,
                                last_name=last_name,
                            )
                            role, _ = role_manager.get_or_create(name=role_key)
                            profile, _ = profile_manager.get_or_create(user_id=user.id)
                            profile.company = company
                            profile.role = role
                            profile.app_scope = app_scope
                            profile.department = department or None
                            if department and role_key not in {RoleType.PLANNER.value, RoleType.STORE.value}:
                                if app_scope not in new_departments:
                                    new_departments[app_scope] = set()
                                new_departments[app_scope].add(department)
                            profile.shift = shift or None
                            profile.phone = phone or None
                            profile.worker_mode_enabled = worker_mode_enabled if role_key == RoleType.SUPERVISOR.value else False
                            profile.planned_shift = planned_shift or None
                            profile.planned_shift_start_date = planned_shift_start_date if planned_shift else None
                            profile.save(using=db_alias)
                            count += 1
                        except Exception as exc:
                            errors.append(f"Row {row_idx}: Error importing employee '{username}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="employee",
                        plural_label="employees",
                        count=count,
                        errors=errors,
                        empty_hint="Add at least one filled employee row to the Excel sheet before uploading.",
                    )
                    if count > 0 and not errors:
                        messages.info(request, "Default imported employee password: Password123!.")

                    if new_departments:
                        settings, _ = SystemSettings.objects.using(db_alias).get_or_create(company=company)
                        catalog = settings.department_catalog or {}
                        catalog_changed = False
                        for scope_key, deps in new_departments.items():
                            existing_deps = catalog.get(scope_key, [])
                            if not isinstance(existing_deps, list):
                                existing_deps = []
                            for d in deps:
                                if not any(e.lower() == d.lower() for e in existing_deps):
                                    existing_deps.append(d)
                                    catalog_changed = True
                            catalog[scope_key] = existing_deps
                        if catalog_changed:
                            settings.department_catalog = catalog
                            settings.save(using=db_alias, update_fields=["department_catalog"])

                    self._mark_onboarding_planner_reset_if_needed(request)

                elif import_type == 'stages':
                    if _reject_if_missing_template_signature(
                        "Production Stages",
                        [
                            ["Name", "Stage Name"],
                            ["Machine Code"],
                            ["Order", "Stage Order", "Is Quality Check", "Color"],
                        ],
                        "Use the Production Stages template and keep Name, Machine Code, Order, and stage settings.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row[0]:
                            continue
                        name = str(row[0]).strip()
                        machine_code = _clean_text(row[1])
                        order = row[2] or 1
                        category = row[3] or None
                        is_qc = row[4] or False
                        color = row[5] or "#90CAF9"

                        if isinstance(is_qc, str):
                            is_qc = is_qc.strip().lower() in {'1', 'true', 'yes', 'y'}
                        else:
                            is_qc = bool(is_qc)

                        machine = _resolve_machine(machine_code)
                        if machine_code and not machine:
                            errors.append(
                                f"Row {row_idx}: Stage '{name}' references unknown machine '{machine_code}'. "
                                "Upload machines first or leave Machine Code blank for a generic routing stage."
                            )
                            continue

                        try:
                            stage_manager.update_or_create(
                                name=name,
                                machine=machine,
                                defaults={
                                    'order': int(order or 1),
                                    'category': category,
                                    'is_quality_check': is_qc,
                                    'color': color or "#90CAF9",
                                }
                            )
                            count += 1
                        except Exception as exc:
                            errors.append(f"Row {row_idx}: Error importing stage '{name}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="stage",
                        plural_label="stages",
                        count=count,
                        errors=errors,
                        empty_hint=(
                            "Add at least one filled stage row. Machine Code is optional; "
                            "when provided it must match an uploaded machine code or name."
                        ),
                    )
                    self._mark_onboarding_planner_reset_if_needed(request)

                elif import_type == 'bom':
                    product_headers = [
                        "Product Name", "Product", "Finished Good", "Finished Product", "Parent Product",
                        "Parent Item", "Assembly", "FG Name", "Item",
                    ]
                    component_headers = [
                        "Component Name", "Component", "Material", "Material Name", "Raw Material",
                        "Part", "Part Name", "Ingredient", "Input Material",
                    ]
                    quantity_headers = [
                        "Component Quantity", "Material Quantity", "Required Quantity", "Required Qty",
                        "Quantity", "Qty", "QTY Per", "Qty Per", "Usage", "Consumption",
                    ]
                    operation_headers = [
                        "Operation Name", "Operation", "Stage", "Stage Name", "Process", "Process Name",
                        "Routing Step", "Work Center Step",
                    ]
                    machine_headers = [
                        "Machine Code", "Machine", "Machine Name", "Resource", "Resource Code",
                        "Work Center", "Work Centre", "Line", "Line Code",
                    ]
                    if _reject_if_missing_template_signature(
                        "BOM",
                        [
                            product_headers,
                            component_headers + operation_headers,
                            quantity_headers + [
                                "Duration (mins)", "Duration", "Duration Minutes",
                                "Run Time", "Run Time (mins)", "Run Minutes",
                                "Minutes Per Unit", "Time Per Unit",
                            ],
                        ],
                        "Use the BOM template and keep Product Name plus component or operation columns.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    grouped_rows = {}
                    invalid_bom_groups = set()
                    carry = {
                        "product": None,
                        "version": None,
                        "status": None,
                        "base_quantity": None,
                        "uom": None,
                        "operation": None,
                        "duration": None,
                        "setup_time": None,
                        "run_time": None,
                        "operation_order": None,
                        "machine_ref": None,
                        "machine_type": None,
                    }
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        explicit_product = _clean_text(_row_value(row, product_headers, 0))
                        if explicit_product and explicit_product.lower() != str(carry["product"] or "").lower():
                            carry.update({
                                "operation": None,
                                "duration": None,
                                "setup_time": None,
                                "run_time": None,
                                "operation_order": None,
                                "machine_ref": None,
                                "machine_type": None,
                            })
                        prod_name = explicit_product or _clean_text(carry["product"])
                        if not prod_name:
                            continue
                        carry["product"] = prod_name

                        component_name = _clean_text(_row_value(row, component_headers, None))
                        operation_name = _clean_text(_first_present(row, operation_headers, None, carry["operation"]))
                        machine_ref = _clean_text(_first_present(row, machine_headers, None, carry["machine_ref"]))
                        machine_type = _clean_text(_first_present(
                            row,
                            ["Machine Type", "Machine Category", "Resource Type", "Work Center Type", "Required Machine Type"],
                            None,
                            carry["machine_type"],
                        ))
                        version_value = _clean_text(_first_present(row, ["Version", "BOM Version", "Revision", "Rev"], None, carry["version"]))
                        status_value = _first_present(row, ["Status", "BOM Status", "State"], None, carry["status"])
                        base_qty_value = _first_present(
                            row,
                            ["Base Quantity", "Base Qty", "Batch Size", "Batch Quantity", "Lot Size", "Output Qty"],
                            None,
                            carry["base_quantity"],
                        )
                        uom_value = _first_present(row, ["BOM UOM", "BOM Unit", "Output Unit", "UOM", "Unit"], None, carry["uom"])
                        duration_value = _first_present(
                            row,
                            ["Duration (mins)", "Duration", "Duration Minutes", "Total Minutes", "Cycle Time", "Cycle Time (mins)"],
                            None,
                            carry["duration"],
                        )
                        setup_value = _first_present(
                            row,
                            ["Setup Time", "Setup Time (mins)", "Setup Minutes", "Setup"],
                            None,
                            carry["setup_time"],
                        )
                        run_value = _first_present(
                            row,
                            ["Run Time", "Run Time (mins)", "Run Minutes", "Minutes Per Unit", "Time Per Unit"],
                            None,
                            carry["run_time"],
                        )
                        order_value = _first_present(
                            row,
                            ["Operation Order", "Stage Order", "Routing Order", "Sequence", "Seq", "Step"],
                            None,
                            carry["operation_order"],
                        )

                        if operation_name:
                            carry["operation"] = operation_name
                        if machine_ref:
                            carry["machine_ref"] = machine_ref
                        if machine_type:
                            carry["machine_type"] = machine_type
                        if version_value:
                            carry["version"] = version_value
                        if status_value is not None and _clean_text(status_value):
                            carry["status"] = status_value
                        if base_qty_value is not None and _clean_text(base_qty_value):
                            carry["base_quantity"] = base_qty_value
                        if uom_value is not None and _clean_text(uom_value):
                            carry["uom"] = uom_value
                        if duration_value is not None and _clean_text(duration_value):
                            carry["duration"] = duration_value
                        if setup_value is not None and _clean_text(setup_value):
                            carry["setup_time"] = setup_value
                        if run_value is not None and _clean_text(run_value):
                            carry["run_time"] = run_value
                        if order_value is not None and _clean_text(order_value):
                            carry["operation_order"] = order_value

                        if not component_name and not operation_name:
                            continue

                        quantity_value = _row_value(row, quantity_headers, None)
                        quantity_decimal = _parse_decimal(quantity_value)
                        if component_name and (quantity_decimal is None or quantity_decimal <= 0):
                            errors.append(
                                f"Row {row_idx}: Component '{component_name}' is missing a valid positive quantity."
                            )
                            invalid_bom_groups.add(prod_name.lower())
                            continue

                        component_unit_value = _row_value(
                            row,
                            ["Component Unit", "Material Unit", "Consumption Unit", "Input Unit", "Unit"],
                            None,
                        )
                        cost_value = _row_value(row, ["Cost Per Unit", "Unit Cost", "Material Cost", "Price"], None)
                        wastage_value = _row_value(row, ["Wastage Quantity", "Wastage Qty", "Scrap Qty", "Waste Qty"], None)
                        scrap_value = _row_value(row, ["Scrap Value Per Unit", "Scrap Value", "Recovery Value"], None)
                        scrap_type_value = _row_value(row, ["Scrap Type", "Wastage Type", "Waste Type"], None)
                        description_value = _row_value(
                            row,
                            ["Description", "Operation Description", "Instructions", "Notes", "QC Trigger"],
                            None,
                        )
                        duration_decimal = _parse_decimal(duration_value)
                        run_decimal = _parse_decimal(run_value)
                        if operation_name and (
                            (duration_decimal is None or duration_decimal <= 0)
                            and (run_decimal is None or run_decimal <= 0)
                        ):
                            errors.append(
                                f"Row {row_idx}: Operation '{operation_name}' is missing Duration or Run Time."
                            )
                            invalid_bom_groups.add(prod_name.lower())
                            continue

                        grouped_rows.setdefault(prod_name.lower(), {"product_name": prod_name, "rows": []})["rows"].append(
                            {
                                "row_idx": row_idx,
                                "component": component_name,
                                "quantity": quantity_decimal if quantity_decimal is not None else Decimal("1"),
                                "component_unit": _normalize_unit(component_unit_value, None),
                                "cost_per_unit": _coerce_decimal(cost_value, "0"),
                                "wastage_quantity": _coerce_decimal(wastage_value, "0"),
                                "scrap_value_per_unit": _coerce_decimal(scrap_value, "0"),
                                "scrap_type": _normalize_scrap_type(scrap_type_value),
                                "operation": operation_name,
                                "duration": _coerce_positive_int(duration_value, 60),
                                "setup_time": max(int(_coerce_decimal(setup_value, "0")), 0),
                                "run_time": _coerce_decimal(run_value, "0"),
                                "operation_order": _coerce_positive_int(order_value, 0) if order_value not in (None, "") else None,
                                "machine_ref": machine_ref,
                                "machine_type": machine_type,
                                "description": _clean_text(description_value),
                                "status": _normalize_status(
                                    status_value,
                                    {"draft", "test", "active"},
                                    "active",
                                ),
                                "version": version_value,
                                "base_quantity": _coerce_decimal(
                                    base_qty_value,
                                    "1",
                                ),
                                "uom": _normalize_unit(uom_value, "pcs"),
                            }
                        )

                    valid_uoms = {"kg", "gm", "m", "cm", "l", "pcs"}
                    for group_key, group in grouped_rows.items():
                        if group_key in invalid_bom_groups:
                            continue
                        product_name = group["product_name"]
                        group_errors = []
                        for item in group["rows"]:
                            if item["machine_ref"] and not _resolve_machine(item["machine_ref"]):
                                group_errors.append(
                                    f"Row {item['row_idx']}: BOM operation '{item['operation'] or 'Unnamed operation'}' references "
                                    f"unknown machine '{item['machine_ref']}'."
                                )
                        if group_errors:
                            errors.extend(group_errors)
                            continue

                        try:
                            with transaction.atomic(using=db_alias):
                                product = product_manager.filter(company=company, name__iexact=product_name).first()
                                if not product:
                                    product = product_manager.create(
                                        name=product_name,
                                        company=company,
                                        unit="pcs",
                                        material_type="finished",
                                    )

                                first_row = group["rows"][0]
                                desired_status = first_row["status"]
                                explicit_version = first_row["version"]
                                version = explicit_version or _next_bom_version(product)
                                existing_bom = bom_manager.filter(product=product, version=version).first()
                                if existing_bom and existing_bom.status == "active":
                                    if explicit_version:
                                        errors.append(
                                            f"Row {first_row['row_idx']}: BOM {version} for '{product_name}' is already active. "
                                            "Use a new version to replace it."
                                        )
                                        continue
                                    version = _next_bom_version(product)
                                    existing_bom = None

                                bom = existing_bom or bom_manager.create(
                                    product=product,
                                    version=version,
                                    status="draft",
                                    created_by_id=request.user.id,
                                )
                                if existing_bom:
                                    bom.status = "draft"
                                bom.base_quantity = first_row["base_quantity"] or Decimal("1")
                                bom.uom = first_row["uom"] if first_row["uom"] in valid_uoms else "pcs"
                                bom.created_by_id = bom.created_by_id or request.user.id
                                bom.save(using=db_alias)

                                bom.components.all().delete()
                                bom.operations.all().delete()

                                operations_by_key = {}
                                order = 1
                                for item in group["rows"]:
                                    component = None
                                    if item["component"]:
                                        material_product = product_manager.filter(
                                            company=company,
                                            name__iexact=item["component"],
                                        ).first()
                                        if not material_product:
                                            material_product = product_manager.create(
                                                company=company,
                                                name=item["component"],
                                                unit=item["component_unit"] if item["component_unit"] in valid_uoms else "pcs",
                                                material_type="raw",
                                            )
                                        elif item["component_unit"] in valid_uoms and material_product.unit != item["component_unit"]:
                                            material_product.unit = item["component_unit"]
                                            material_product.save(using=db_alias, update_fields=["unit"])
                                        component = bom_component_manager.create(
                                            bom=bom,
                                            product=material_product,
                                            material_name=item["component"],
                                            quantity=item["quantity"] or Decimal("1"),
                                            unit=material_product.unit or "pcs",
                                            cost_per_unit=item["cost_per_unit"],
                                            wastage_quantity=item["wastage_quantity"],
                                            scrap_value_per_unit=item["scrap_value_per_unit"],
                                            scrap_type=item["scrap_type"],
                                        )

                                    if item["operation"]:
                                        machine = _resolve_machine(item["machine_ref"])
                                        if item["machine_ref"] and not machine:
                                            errors.append(
                                                f"Row {item['row_idx']}: BOM operation '{item['operation']}' references "
                                                f"unknown machine '{item['machine_ref']}'."
                                            )
                                            continue
                                        machine_type = item["machine_type"] or ((machine.category or machine.type) if machine else None)
                                        operation_order = item["operation_order"] or order
                                        setup_time = item["setup_time"]
                                        run_time = item["run_time"] if item["run_time"] > 0 else Decimal(item["duration"])
                                        duration_minutes = item["duration"]
                                        if item["run_time"] > 0 and setup_time > 0:
                                            duration_minutes = max(
                                                int((item["run_time"] * (first_row["base_quantity"] or Decimal("1"))) + setup_time),
                                                1,
                                            )
                                        stage, _ = stage_manager.get_or_create(
                                            name=item["operation"],
                                            machine=machine,
                                            defaults={
                                                "order": operation_order,
                                                "category": machine_type,
                                            },
                                        )
                                        operation_key = (
                                            item["operation"].lower(),
                                            machine.id if machine else None,
                                            operation_order,
                                        )
                                        operation = operations_by_key.get(operation_key)
                                        if not operation:
                                            operation = bom_operation_manager.create(
                                                bom=bom,
                                                machine=machine,
                                                stage=stage,
                                                duration_minutes=duration_minutes,
                                                run_time=run_time,
                                                setup_time=setup_time,
                                                order=operation_order,
                                                machine_type=machine_type,
                                                description=item["description"],
                                            )
                                            operations_by_key[operation_key] = operation
                                            order = max(order, operation_order + 1)
                                        if component:
                                            bom_operation_material_manager.get_or_create(
                                                operation=operation,
                                                component=component,
                                            )

                                if desired_status == "active":
                                    previous_active = bom_manager.filter(
                                        product=product,
                                        status="active",
                                    ).exclude(id=bom.id).first()
                                    bom.status = "active"
                                    bom.save(using=db_alias, update_fields=["status"])
                                    if previous_active:
                                        flag_bom_change_impact(bom, actor=request.user)
                                else:
                                    bom.status = desired_status
                                    bom.save(using=db_alias, update_fields=["status"])

                                count += len(group["rows"])
                        except Exception as exc:
                            errors.append(f"Error importing BOM for '{product_name}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="BOM row",
                        plural_label="BOM rows",
                        count=count,
                        errors=errors,
                        empty_hint="Add at least one filled BOM row to the Excel sheet before uploading.",
                    )
                    self._mark_onboarding_planner_reset_if_needed(request)

                elif import_type == 'work_orders':
                    if _reject_if_missing_template_signature(
                        "Work Orders",
                        [
                            ["Product Name", "Product"],
                            ["Quantity", "Qty", "WO Qty"],
                            ["Start Date (YYYY-MM-DD)", "Start Date", "Scheduled Start", "End Date (YYYY-MM-DD)", "End Date", "Due Date"],
                        ],
                        "Use the Work Orders template and keep Product Name, Quantity, and planned date columns.",
                    ):
                        return self._redirect_to_next(request, 'bulk_import_dashboard')

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        product_name = _clean_text(_row_value(row, ["Product Name", "Product"], 0))
                        if not product_name:
                            continue

                        qty = _row_value(row, ["Quantity", "Qty", "WO Qty"], 1)
                        start = _row_value(row, ["Start Date (YYYY-MM-DD)", "Start Date", "Scheduled Start"], 2)
                        end = _row_value(row, ["End Date (YYYY-MM-DD)", "End Date", "Due Date"], 3)
                        status = _row_value(row, ["Status (pending/in_progress)", "Status", "WO Status"], 4)
                        email = _clean_text(_row_value(row, ["Assigned To (Email)", "Assigned To", "Email"], 5))
                        priority = _normalize_status(
                            _row_value(row, ["Priority"], 6),
                            {"low", "normal", "high", "urgent"},
                            "normal",
                        ).capitalize()

                        assigned_user = None
                        if email:
                            assigned_user = user_manager.filter(
                                Q(profile__company=company),
                                Q(username=email) | Q(email=email)
                            ).first()

                        start_dt = _coerce_datetime(start)
                        end_dt = _coerce_datetime(end)
                        requested_status = _normalize_status(status, {"pending"}, "pending")
                        raw_status = _clean_text(status).lower()
                        if raw_status == "draft":
                            errors.append(
                                f"Row {row_idx}: Work order status 'draft' was imported as pending because draft work orders are no longer used."
                            )
                        elif raw_status and raw_status not in {"pending"}:
                            errors.append(
                                f"Row {row_idx}: Work order status '{raw_status}' was imported as pending. "
                                "Bulk upload cannot bypass material, machine, and worker gates."
                            )

                        product = product_manager.filter(company=company, name__iexact=product_name).first()
                        linked_bom = None
                        if product:
                            linked_bom = (
                                bom_manager.filter(product=product, status='active')
                                .order_by('-created_at', '-id')
                                .first()
                            )
                        if not linked_bom:
                            errors.append(
                                f"Row {row_idx}: No active BOM found for '{product_name}'. "
                                "Upload and activate the BOM before importing work orders."
                            )
                            continue

                        try:
                            wo = work_order_manager.create(
                                company=company,
                                product_name=product_name,
                                bom=linked_bom,
                                quantity=_coerce_positive_int(qty, 1),
                                scheduled_start_date=start_dt,
                                start_date=start_dt,
                                end_date=end_dt,
                                due_date=end_dt,
                                status=requested_status,
                                assigned_to=assigned_user or request.user,
                                priority=priority,
                                operation_flow_mode=get_company_default_operation_flow_mode(company),
                            )
                            NotificationService.notify_role(
                                company,
                                roles=["store", "admin"],
                                title="Material check requested",
                                message=f"WO #{wo.id} needs BOM material readiness for {wo.quantity} units.",
                                link="/manufacturing/store/",
                                exclude_user=request.user,
                            )
                            count += 1
                        except Exception as exc:
                            errors.append(f"Row {row_idx}: Error importing work order '{product_name}': {exc}")

                    self._notify_import_result(
                        request,
                        singular_label="work order",
                        plural_label="work orders",
                        count=count,
                        errors=errors,
                        empty_hint="Add at least one filled work-order row to the Excel sheet before uploading.",
                    )
                    self._mark_onboarding_planner_reset_if_needed(request)

                else:
                    messages.error(request, "Unsupported import type.")

            except Exception as exc:
                messages.error(request, f"Import Failed: {exc}")

            return self._redirect_to_next(request, 'bulk_import_dashboard')
        finally:
            reset_current_tenant_db(ctx_token)

