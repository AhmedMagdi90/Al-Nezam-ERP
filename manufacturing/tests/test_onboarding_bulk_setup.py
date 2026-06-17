from django.test import RequestFactory, SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from manufacturing.import_catalog import get_bulk_import_catalog, get_bulk_import_filenames
from manufacturing.views.bulk import HandleBulkImportView


class OnboardingBulkSetupTests(SimpleTestCase):
    def test_import_catalog_contains_all_supported_setup_types(self):
        catalog = get_bulk_import_catalog()
        import_types = {item["type"] for item in catalog}

        self.assertEqual(
            import_types,
            {"machines", "products", "stages", "bom", "work_orders", "employees"},
        )
        self.assertTrue(all(item["template"].endswith(".xlsx") for item in catalog))
        self.assertTrue(all(item["sample"].endswith(".xlsx") for item in catalog))

    def test_downloadable_filenames_are_built_from_catalog(self):
        filenames = get_bulk_import_filenames()

        self.assertIn("machines_template.xlsx", filenames)
        self.assertIn("employees_template.xlsx", filenames)
        self.assertIn("bom_sample.xlsx", filenames)
        self.assertIn("work_orders_sample.xlsx", filenames)

    def test_employee_template_matches_current_team_fields(self):
        workbook = load_workbook("templates/downloads/employees_template.xlsx")
        sheet = workbook["Employees"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        self.assertEqual(workbook.sheetnames, ["Employees", "Instructions", "Choices"])
        self.assertEqual(
            headers,
            [
                "Employee Name",
                "Email",
                "Phone",
                "Role",
                "Department(s)",
                "Shift",
                "App Scope",
                "Worker Mode Enabled",
                "Initial Password",
            ],
        )
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:K501")

    def test_employee_template_has_dropdowns_for_fixed_choices(self):
        workbook = load_workbook("templates/downloads/employees_template.xlsx")
        sheet = workbook["Employees"]
        choices = workbook["Choices"]
        validation_ranges = {
            str(cell_range)
            for validation in sheet.data_validations.dataValidation
            for cell_range in validation.cells.ranges
        }
        app_scope_validations = [
            validation
            for validation in sheet.data_validations.dataValidation
            if "G2:G501" in {str(cell_range) for cell_range in validation.cells.ranges}
        ]

        self.assertIn("D2:D501", validation_ranges)
        self.assertIn("F2:F501", validation_ranges)
        self.assertIn("G2:G501", validation_ranges)
        self.assertIn("H2:H501", validation_ranges)
        self.assertEqual(len(app_scope_validations), 1)
        self.assertIn("manufacturing", app_scope_validations[0].formula1)
        self.assertNotIn("planner", app_scope_validations[0].formula1)
        self.assertEqual(choices["C2"].value, "manufacturing")
        self.assertNotIn(
            "planner",
            [choices.cell(row=row, column=3).value for row in range(2, 8)],
        )

    def test_bom_template_matches_reworked_import_fields(self):
        workbook = load_workbook("templates/downloads/bom_template.xlsx")
        sheet = workbook["BOM Rows"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        validation_ranges = {
            str(cell_range)
            for validation in sheet.data_validations.dataValidation
            for cell_range in validation.cells.ranges
        }

        self.assertEqual(workbook.sheetnames, ["BOM Rows", "Choices", "Instructions"])
        self.assertEqual(
            headers,
            [
                "Product Name",
                "BOM Version",
                "BOM Status",
                "Base Quantity",
                "BOM UOM",
                "Operation Order",
                "Operation Name",
                "Machine Code",
                "Machine Type",
                "Setup Time (mins)",
                "Run Time (mins/unit)",
                "Duration (mins)",
                "Component Name",
                "Component Quantity",
                "Component Unit",
                "Cost Per Unit",
                "Wastage Quantity",
                "Scrap Value Per Unit",
                "Scrap Type",
                "Instructions",
            ],
        )
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:T501")
        self.assertIn("C2:C501", validation_ranges)
        self.assertIn("E2:E501", validation_ranges)
        self.assertIn("O2:O501", validation_ranges)
        self.assertIn("S2:S501", validation_ranges)

    def test_onboarding_data_template_uses_shared_import_catalog(self):
        with open("templates/registration/onboarding_data.html", encoding="utf-8") as handle:
            content = handle.read()

        self.assertIn("{% for card in import_catalog %}", content)
        self.assertIn('name="import_type" value="{{ card.type }}"', content)
        self.assertIn('name="next" value="onboarding_data"', content)
        self.assertIn('name="tenant_code" value="{{ active_tenant_code }}"', content)
        self.assertIn("{% url 'download_template' card.template %}", content)
        self.assertIn('data-upload-form="true"', content)
        self.assertIn('data-upload-input="true"', content)

    def test_onboarding_users_template_preserves_tenant_code_for_employee_upload(self):
        with open("templates/registration/onboarding_users.html", encoding="utf-8") as handle:
            content = handle.read()

        self.assertIn('name="import_type" value="employees"', content)
        self.assertIn('name="tenant_code" value="{{ active_tenant_code }}"', content)
        self.assertIn('name="next" value="onboarding_users"', content)
        self.assertIn("Skip for Now", content)

    def test_bulk_import_redirect_helper_honors_named_next_target(self):
        request = RequestFactory().post("/manufacturing/bulk-import/upload/", {"next": "onboarding_data"})

        response = HandleBulkImportView._redirect_to_next(request, "bulk_import_dashboard")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("onboarding_data"))

    def test_bulk_import_redirect_helper_falls_back_when_next_missing(self):
        request = RequestFactory().post("/manufacturing/bulk-import/upload/")

        response = HandleBulkImportView._redirect_to_next(request, "bulk_import_dashboard")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("bulk_import_dashboard"))
