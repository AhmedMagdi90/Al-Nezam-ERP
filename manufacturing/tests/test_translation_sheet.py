from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import translation
from openpyxl import Workbook, load_workbook

from manufacturing.runtime_translations import get_company_translation_map, resolve_runtime_translation
from manufacturing.tests.utils import create_company, create_user_with_role


class TranslationSheetTests(TestCase):
    def setUp(self):
        self.company = create_company("Sheet Co")
        self.admin = create_user_with_role("sheet_admin", "admin", self.company)
        self.client.force_login(self.admin)
        self.url = reverse("settings_dashboard")

    def _build_workbook(self, rows):
        workbook = Workbook()
        workbook.active.title = "Instructions"
        sheet = workbook.create_sheet("Translations")
        sheet.append([
            "English Source",
            "Arabic Translation",
            "Screen Hint",
            "Template Location",
            "Status",
        ])
        for row in rows:
            sheet.append(row)

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return payload

    def test_download_translation_sheet_returns_expected_headers(self):
        response = self.client.get(
            f"{self.url}?tab=translations&scope=planner&action=download_translation_sheet"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Translations", workbook.sheetnames)
        sheet = workbook["Translations"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                "English Source",
                "Arabic Translation",
                "Screen Hint",
                "Template Location",
                "Status",
            ],
        )

        rows = {
            str(row[0]): row
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertIn("Settings", rows)
        self.assertTrue(rows["Settings"][2])

    def test_upload_translation_sheet_applies_company_overrides(self):
        workbook = self._build_workbook(
            [
                [
                    "Actual vs Planned",
                    "فعلي مقابل المخطط",
                    "Reports Dashboard",
                    "reports_dashboard.html",
                    "global",
                ]
            ]
        )
        upload = SimpleUploadedFile(
            "translations.xlsx",
            workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            self.url,
            data={
                "action": "import_translation_sheet",
                "scope": "planner",
                "translation_sheet": upload,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            get_company_translation_map(self.company, "ar").get("Actual vs Planned"),
            "فعلي مقابل المخطط",
        )

        with translation.override("ar"):
            self.assertEqual(
                resolve_runtime_translation(self.company, "Actual vs Planned"),
                "فعلي مقابل المخطط",
            )

    def test_upload_translation_sheet_clears_existing_company_override_when_arabic_is_blank(self):
        workbook = self._build_workbook(
            [
                [
                    "Actual vs Planned",
                    "فعلي مقابل المخطط",
                    "Reports Dashboard",
                    "reports_dashboard.html",
                    "company",
                ]
            ]
        )
        first_upload = SimpleUploadedFile(
            "translations.xlsx",
            workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.post(
            self.url,
            data={
                "action": "import_translation_sheet",
                "scope": "planner",
                "translation_sheet": first_upload,
            },
            follow=True,
        )

        clear_workbook = self._build_workbook(
            [
                [
                    "Actual vs Planned",
                    "",
                    "Reports Dashboard",
                    "reports_dashboard.html",
                    "company",
                ]
            ]
        )
        clear_upload = SimpleUploadedFile(
            "translations-clear.xlsx",
            clear_workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            self.url,
            data={
                "action": "import_translation_sheet",
                "scope": "planner",
                "translation_sheet": clear_upload,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("Actual vs Planned", get_company_translation_map(self.company, "ar"))
