"""
Manufacturing Views Module

This module handles the core business logic and HTTP responses for the Manufacturing app.
It is currently monolithic and contains:
1. Dashboard Views (Planner, Supervisor, Factory Setup)
2. AJAX/API Endpoints (Timeline data, Work Order updates)
3. Form Handling (Work Order creation, BOM Builder)

Future Refactoring Recommendation:
- Split into `views/dashboard.py`, `views/api.py`, and `views/forms.py`.
- Move complex logic to `services.py`.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.contrib import messages
import json

from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.contrib.auth.models import User

from accounts.models import Profile
from .forms import WorkOrderForm
from .models import (
    WorkOrder,
    Machine,
    ProductionStage,
    BillOfMaterial,
    BOMComponent,
    BOMOperation,
    Product,
    ProductionLog,
    MachineFault,
)
from manufacturing.utils import normalize_operation_time_minutes
from .forms import ProductionLogForm
from decimal import Decimal
from datetime import timedelta
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Sum

# ---------------------------------------------------------------------
# 🧠 Utility: Role-based Access & Company Isolation
# ---------------------------------------------------------------------
def user_has_role(user, allowed_roles):
    """Check if a user's role matches any allowed role (case-insensitive)."""
    if user.is_superuser:
        return True
    try:
        return user.profile.role.name.lower() in [r.lower() for r in allowed_roles]
    except Exception:
        return False

def get_user_company(user):
    """Safely get user's company, returns None if not found."""
    try:
        if hasattr(user, 'profile') and user.profile.company:
            return user.profile.company
    except Exception:
        pass
    return None

def require_company(user):
    """Raise error if user doesn't have a company."""
    company = get_user_company(user)
    if not company:
        raise ValueError("User must have a company assigned.")
    return company

def landing_page(request):
    """Public Landing Page."""
    if request.user.is_authenticated:
        return redirect('dashboard') 
    return render(request, 'landing.html')


# ---------------------------------------------------------------------
# 🚀 Dashboard (Stub for role dispatch)
# ---------------------------------------------------------------------

@login_required
def dashboard(request):
    """
    Central Dashboard Dispatcher.
    Redirects users based on their role.
    """
    user = request.user
    if not hasattr(user, 'profile'):
        return redirect('landing_page')
        
    role = user.profile.role.name.lower()
    
    if role == 'admin' or role == 'owner':
        return redirect('supervisor_dashboard') # Or reports_dashboard
    elif role == 'planner':
        return redirect('planner_dashboard')
    elif role == 'supervisor':
        return redirect('supervisor_dashboard')
    elif role == 'worker':
        return redirect('record_output')
    elif role == 'quality':
        return redirect('quality_check')
    elif role == 'maintenance':
        return redirect('maintenance_dashboard')
    else:
        return redirect('landing_page')
# NOTE: This should eventually replace the simple redirect in 'dashboard'
# But 'register_company' is for SaaS Onboarding.

def register_company(request):
    """
    Step 1: SaaS Registration
    - Create Company
    - Create Owner User
    - Mock Payment
    """
    from .forms import CompanyRegistrationForm
    from .models import Company
    from django.contrib.auth import login
    from accounts.models import Role # Need Role model for Owner
    
    if request.method == "POST":
        form = CompanyRegistrationForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            # 1. Create User
            if User.objects.filter(username=data['owner_email']).exists():
                messages.error(request, "Email already registered.")
                return render(request, 'registration/register_company.html', {'form': form})
                
            user = User.objects.create_user(
                username=data['owner_email'],
                email=data['owner_email'],
                password=data['owner_password']
            )
            
            # 2. Create Company
            company = Company.objects.create(
                name=data['company_name'],
                subscription_plan='free_trial'
            )
            
            # 3. Create Profile & Role
            # Ensure 'admin' role (Owner) exists
            owner_role, _ = Role.objects.get_or_create(name='admin')
            
            # Accounts.Profile handles the link. 
            # Note: accounts.signals might auto-create profile, so we check first
            if hasattr(user, 'profile'):
                profile = user.profile
                profile.company = company
                profile.role = owner_role
                profile.save()
            else:
                Profile.objects.create(user=user, role=owner_role, company=company)
                
            # 4. Login & Redirect
            login(request, user)
            messages.success(request, f"Welcome to Nezam! Your company {company.name} is ready.")
            return redirect('onboarding_data') # Next Step
            
    else:
        form = CompanyRegistrationForm()
        
    return render(request, 'registration/register_company.html', {'form': form}) 


@login_required
def onboarding_data(request):
    """
    Step 2: Load Data
    - Option A: One-click "Load Demo Factory"
    - Option B: Bulk Upload (uses existing logic)
    """
    if request.method == "POST":
        if 'load_demo' in request.POST:
            # 🔮 Load Standard Demo Data (Products, Machines, BOMs)
            from .models import Machine, Product
            company = request.user.profile.company
            
            # TODO: SaaS Isolation (In future, validade company ownership of objects)
            
            if not Machine.objects.filter(company=company, code="A001").exists():
                Machine.objects.create(company=company, name="Assembly Line 1", code="A001", type="Assembly")
                
            if not Product.objects.filter(company=company, name="Demo Product A").exists():
                Product.objects.create(company=company, name="Demo Product A", material_type="finished")
            
            messages.success(request, "✅ Demo Factory Loaded!")
            return redirect('onboarding_users') # Next Step
            
        elif request.FILES:
             # Redirect to handle_bulk_import logic explicitly or reuse it
             # Simplest way: Call handle_bulk_import internally or redirect POST?
             # Redirect POST is hard. Let's redirect user to Bulk Import Page or just call logic.
             # Better: Allow handle_bulk_import to be called directly if we refactor.
             # For now: Just tell user to use valid endpoint or fix HTML form to point to handle_bulk_import!
             # Wait, editing View is safer.
             return handle_bulk_import(request)
            
    return render(request, 'registration/onboarding_data.html') 


@login_required
def onboarding_users(request):
    """
    Step 3: Setup Team
    - Add basic users for key roles.
    """
    from accounts.models import Role, Profile
    
    if request.method == "POST":
        company = request.user.profile.company
        
        # Iterate over roles to see who was added
        roles_to_add = ['planner', 'supervisor', 'worker', 'quality', 'maintenance']
        
        created_count = 0
        for role_key in roles_to_add:
            email = request.POST.get(f'email_{role_key}')
            name = request.POST.get(f'name_{role_key}')
            
            if email and name:
                if not User.objects.filter(username=email).exists():
                    # Create User
                    # Default password: "Password123!" (In real app, send invite email)
                    u = User.objects.create_user(username=email, email=email, password="Password123!", first_name=name)
                    r, _ = Role.objects.get_or_create(name=role_key)
                    
                    # Create/Update Profile
                    if hasattr(u, 'profile'):
                         p = u.profile
                         p.role = r
                         p.company = company
                         p.save()
                    else:
                         Profile.objects.create(user=u, role=r, company=company)
                    
                    created_count += 1
        
        messages.success(request, f"🎉 Setup Complete! {created_count} team members added.")
        return redirect('dashboard') # DONE!
        
    return render(request, 'registration/onboarding_users.html') 



# ---------------------------------------------------------------------
# 🧾 Create Work Order (Planner / Admin)
# ---------------------------------------------------------------------
@login_required
def create_work_order(request):
    """
    Deprecated: Redirects to Main Planner Dashboard.
    Kept for legacy URL compatibility.
    """
    return redirect('planner_dashboard')


# ---------------------------------------------------------------------
# 📅 Planner Dashboard (Timeline & Command Center)
# ---------------------------------------------------------------------
@login_required
def planner_dashboard(request):
    """
    Main Manufacturing Dashboard (The "App Shell" View).
    Consolidates Timeline, Work Orders, and Machine Management.
    """
    if not user_has_role(request.user, ['planner', 'admin', 'supervisor']):
        return HttpResponseForbidden("🚫 You are not authorized to access the planner dashboard.")

    # --- AJAX requests ---
    ajax_type = request.GET.get("ajax")
    status_filter = request.GET.get("status", "")
    search = request.GET.get("search", "")

    company = get_user_company(request.user)
    if not company:
        return JsonResponse({"success": False, "error": "Company not found"}, status=403)
    
    if ajax_type == "workorders":
        work_orders = WorkOrder.objects.filter(company=company).select_related("machine", "bom", "assigned_to").order_by("-id")
        if status_filter:
            work_orders = work_orders.filter(status=status_filter)
        if search:
            work_orders = work_orders.filter(product_name__icontains=search)

        html = render_to_string("manufacturing/partials/workorders_list.html", {"work_orders": work_orders})
        return JsonResponse({"success": True, "html": html})

    elif ajax_type == "boms":
        boms = BillOfMaterial.objects.filter(product__company=company).select_related("product", "created_by").order_by("-created_at")
        html = render_to_string("manufacturing/partials/boms_list.html", {"boms": boms})
        return JsonResponse({"success": True, "html": html})

    elif ajax_type == "stages":
        stages = ProductionStage.objects.filter(machine__company=company).select_related("machine").order_by("order")
        html = render_to_string("manufacturing/partials/stages_list.html", {"stages": stages})
        return JsonResponse({"success": True, "html": html})
    
    elif ajax_type == "timeline":
        import json
        from django.core.serializers.json import DjangoJSONEncoder
        
        machines = list(Machine.objects.filter(company=company)) 
        work_orders = WorkOrder.objects.filter(company=company).select_related("machine").order_by("start_date")
        
        active_wos = {wo.machine_id: wo for wo in work_orders if wo.status == 'in_progress' and wo.machine_id}
        for m in machines:
             m.active_wo = active_wos.get(m.id)

        machines_data = [
            {"id": m.id, "name": m.name, "status": m.status} 
            for m in machines
        ]
        
        tasks_data = [
            {
                "id": wo.id,
                "machine_id": wo.machine.id if wo.machine else None,
                "product": wo.product_name,
                "start": wo.start_date.isoformat() if wo.start_date else None,
                "end": wo.end_date.isoformat() if wo.end_date else None,
                "status": wo.status,
                "progress": getattr(wo, 'progress', 0)
            }
            for wo in work_orders if wo.machine
        ]
        
        hours = list(range(24))  # 0-23 hours
        
        html = render_to_string("manufacturing/partials/timeline.html", {
            "machines": machines, 
            "work_orders": work_orders,
            "pending_wos": WorkOrder.objects.filter(company=company, status='pending', machine__isnull=True),
            "machines_json": json.dumps(machines_data, cls=DjangoJSONEncoder),
            "tasks_json": json.dumps(tasks_data, cls=DjangoJSONEncoder),
            "hours": hours,
        })
        return JsonResponse({"success": True, "html": html})


    # --- Regular GET/POST ---
    company = require_company(request.user)
    
    # 🎯 Optimize Timeline Loading: Pre-fetch data for "Zero-Fetch" rendering
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    
    tl_machines = list(Machine.objects.filter(company=company))
    tl_wos = WorkOrder.objects.filter(company=company).select_related("machine")
    
    # Active Machine States
    tl_active_wos = {wo.machine_id: wo for wo in tl_wos if wo.status == 'in_progress' and wo.machine_id}
    for m in tl_machines:
         m.active_wo = tl_active_wos.get(m.id)

    tl_machines_data = [{"id": m.id, "name": m.name, "status": m.status} for m in tl_machines]
    
    tl_tasks_data = [
        {
            "id": wo.id,
            "machine_id": wo.machine.id if wo.machine else None,
            "product": wo.product_name,
            "start": wo.start_date.isoformat() if wo.start_date else None,
            "end": wo.end_date.isoformat() if wo.end_date else None,
            "start": wo.start_date.isoformat() if wo.start_date else None,
            "end": wo.end_date.isoformat() if wo.end_date else None,
            "status": wo.status,
            "progress": getattr(wo, 'progress', 0)
        }
        for wo in tl_wos if wo.machine
    ]

    # --- 🆕 NEW DASHBOARD DATA ---
    # 1. Recent BOMs
    recent_boms = BillOfMaterial.objects.filter(product__company=company).select_related('product').order_by('-updated_at')[:5]

    # 2. Active Work Orders (for list widget)
    active_dashboard_wos = WorkOrder.objects.filter(company=company, status='in_progress').order_by('start_date')[:5]

    # 3. Machine Overview Stats
    # Simple calculation based on status string if available, otherwise assume all operational or active
    op_count = sum(1 for m in tl_machines if m.status == 'operational' or m.active_wo)
    fault_count = sum(1 for m in tl_machines if m.status == 'maintenance')
    free_count = len(tl_machines) - op_count - fault_count
    
    # 4. Pending Tasks (Mock/Real Mix)
    pending_tasks_count = WorkOrder.objects.filter(company=company, status='pending').count()

    common_context = {
        "machines": tl_machines,
        "machines_json": json.dumps(tl_machines_data, cls=DjangoJSONEncoder),
        "tasks_json": json.dumps(tl_tasks_data, cls=DjangoJSONEncoder),
        "pending_wos": WorkOrder.objects.filter(company=company, status='pending', machine__isnull=True),
        "work_orders": WorkOrder.objects.filter(company=company).order_by("-id")[:50], # Recent WOs
        "stages": ProductionStage.objects.filter(machine__company=company).order_by("order"),
        "boms": BillOfMaterial.objects.filter(product__company=company).order_by("-id"),
        # New Keys
        "recent_boms": recent_boms,
        "active_dashboard_wos": active_dashboard_wos,
        "machine_stats": [op_count, fault_count, free_count],
        "pending_tasks_count": pending_tasks_count
    }
    
    if request.method == "POST":
        form = WorkOrderForm(request.POST)
        form.fields['bom'].queryset = BillOfMaterial.objects.filter(product__company=company)
        
        if form.is_valid():
            work_order = form.save(commit=False)
            if work_order.bom and work_order.bom.product.company != company:
                messages.error(request, "❌ Security Error: BOM does not belong to your company.")
                return redirect("planner_dashboard")
            
            if work_order.bom:
                work_order.product_name = work_order.bom.product.name
            else:
                work_order.product_name = "Custom Order"

            work_order.assigned_to = request.user
            work_order.company = company
            work_order.save()

            from .services import WorkOrderService
            WorkOrderService.create_subtasks(work_order, request.user, company)
            
            messages.success(request, "✅ Work order and sub-tasks created successfully!")
            return redirect("planner_dashboard")
        else:
            messages.error(request, "❌ Please fix the errors in the form.")
    else:
        form = WorkOrderForm()
        form.fields['bom'].queryset = BillOfMaterial.objects.filter(product__company=company)

    common_context['form'] = form
    # We render the NEW unified template:
    return render(request, "manufacturing/planner_dashboard.html", common_context)


# ---------------------------------------------------------------------
# 🏭 Factory Setup Dashboard
# ---------------------------------------------------------------------
@login_required
def factory_setup(request):
    """
    Dedicated Factory Setup page for managing machines, stages, and BOMs.
    Provides a centralized interface for factory configuration.
    """
    if not user_has_role(request.user, ['planner', 'admin', 'supervisor']):
        return HttpResponseForbidden("🚫 You are not authorized to access factory setup.")
    
    company = require_company(request.user)
    
    # Get all machines, stages, and BOMs for display
    machines = Machine.objects.filter(company=company).order_by('name')
    stages = ProductionStage.objects.filter(machine__company=company).select_related('machine').order_by('order')
    boms = BillOfMaterial.objects.filter(product__company=company).select_related('product').order_by('-created_at')[:10]
    
    # Get active products for BOM modal
    active_products = Product.objects.filter(company=company)
    
    context = {
        'machines': machines,
        'stages': stages,
        'boms': boms,
        'active_products': active_products,
    }
    
    return render(request, 'manufacturing/factory_setup.html', context)




# ---------------------------------------------------------------------
# 🔧 Update Work Order (AJAX)
# ---------------------------------------------------------------------
@require_POST
@login_required
def update_work_order_status(request):
    """Update status or assigned operator of a work order via AJAX."""
    try:
        company = require_company(request.user)
        wo_id = request.POST.get("id")
        status = request.POST.get("status")
        assigned_to_id = request.POST.get("assigned_to")

        wo = WorkOrder.objects.filter(company=company).get(id=wo_id)
        
        # Track if machine changed (for rescheduling)
        machine_changed = False
        
        # Update fields if provided
        if status:
            wo.status = status
        if assigned_to_id:
            # Ensure assigned user belongs to the same company
            assigned_user = User.objects.filter(profile__company=company).get(id=assigned_to_id)
            wo.assigned_to = assigned_user
            
        stage_id = request.POST.get("stage_id")
        if stage_id:
            wo.current_stage_id = stage_id

        # 📅 Scheduling Updates (Drag & Drop)
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        machine_id = request.POST.get("machine_id")

        if start_date:
            wo.start_date = start_date
        if end_date:
            wo.end_date = end_date
        if machine_id:
            # Check if machine actually changed
            if str(wo.machine_id) != str(machine_id):
                machine_changed = True
                # Validate machine belongs to company
                from .models import Machine
                Machine.objects.filter(company=company, id=machine_id).get()  # Raises DoesNotExist if invalid
                wo.machine_id = machine_id
            
        wo.save()

        # 🔄 Reschedule subtasks if machine changed and user is planner/admin
        if machine_changed and user_has_role(request.user, ['planner', 'admin']):
            from .services import WorkOrderService
            # If this is a subtask, reschedule all subtasks of its parent
            if wo.parent:
                WorkOrderService.reschedule_subtasks(wo.parent)
            # If this is a parent WO, reschedule all its subtasks
            elif wo.sub_tasks.exists():
                WorkOrderService.reschedule_subtasks(wo)

        return JsonResponse({"success": True, "rescheduled": machine_changed})
    except WorkOrder.DoesNotExist:
        return JsonResponse({"success": False, "error": "Work order not found."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ---------------------------------------------------------------------
# 🔮 Simulation & Live Costing API
# ---------------------------------------------------------------------
@login_required
@require_POST
def simulate_work_order(request):
    """
    API to simulate a production run before committing.
    """
    try:
        data = json.loads(request.body)
        bom_id = data.get("bom_id")
        quantity = int(data.get("quantity", 1))
        
        if not bom_id:
             return JsonResponse({"success": False, "error": "BOM ID required"})

        company = require_company(request.user)
        bom = BillOfMaterial.objects.filter(product__company=company).get(id=bom_id)
        
        # Run Simulation
        from .services import BomService # Note: capital S for consistency if renamed, assuming BOMService in file
        from .services import BOMService # importing correct name
        
        result = BOMService.simulate_run(bom, quantity)
        
        return JsonResponse({"success": True, "data": result})
        
    except BillOfMaterial.DoesNotExist:
        return JsonResponse({"success": False, "error": "BOM not found"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
@require_POST
def analyze_quality_image(request):
    """
    Simulate processing an image for defects.
    """
    try:
        if 'image' not in request.FILES:
             return JsonResponse({"success": False, "error": "No image uploaded"})
             
        image = request.FILES['image']
        
        from .services import QualityService
        result = QualityService.analyze_image(image)
        
        return JsonResponse({"success": True, "data": result})
        
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ---------------------------------------------------------------------
# ⚙️ Create Machine (AJAX)
# ---------------------------------------------------------------------
@login_required
@require_POST
def create_machine(request):
    """Create a new Machine for the planner's company."""
    if not user_has_role(request.user, ['planner', 'admin']):
        return JsonResponse({"success": False, "error": "Unauthorized."}, status=403)
    
    try:
        company = require_company(request.user)
        name = request.POST.get("name")
        code = request.POST.get("code")
        machine_type = request.POST.get("type", "")
        status = request.POST.get("status", "operational")

        if not name or not code:
            return JsonResponse({"success": False, "error": "Name and Code are required."}, status=400)

        # Check if code already exists for this company
        if Machine.objects.filter(company=company, code=code).exists():
            return JsonResponse({"success": False, "error": f"Machine with code '{code}' already exists."}, status=400)

        machine = Machine.objects.create(
            company=company,
            name=name,
            code=code,
            type=machine_type,
            status=status,
            is_active=True
        )
        return JsonResponse({"success": True, "id": machine.id, "message": f"Machine '{name}' created successfully!"})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ---------------------------------------------------------------------
# ⚙️ Create Stage (AJAX)
# ---------------------------------------------------------------------
@login_required
@require_POST
def create_stage(request):
    """Create a new Production Stage linked to a specific machine."""
    if not user_has_role(request.user, ['planner', 'admin']):
        return JsonResponse({"success": False, "error": "Unauthorized."}, status=403)
    
    try:
        company = require_company(request.user)
        name = request.POST.get("name")
        machine_id = request.POST.get("machine")
        order = int(request.POST.get("order", 1))
        color = request.POST.get("color", "#3B82F6")
        is_quality_check = request.POST.get("is_quality_check") == "true"

        if not name or not machine_id:
            return JsonResponse({"success": False, "error": "Name and Machine are required."}, status=400)

        machine = Machine.objects.filter(company=company).get(id=machine_id)
        stage = ProductionStage.objects.create(
            name=name, 
            machine=machine, 
            order=order, 
            color=color,
            is_quality_check=is_quality_check
        )
        return JsonResponse({"success": True, "id": stage.id, "message": f"Stage '{name}' created successfully!"})

    except Machine.DoesNotExist:
        return JsonResponse({"success": False, "error": "Machine not found."}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)



# ---------------------------------------------------------------------
# 🧾 BOM Builder (Full Page, 3 Tabs)
# ---------------------------------------------------------------------
@login_required
def bom_builder(request, bom_id=None):
    company = get_user_company(request)
    if not company: return redirect('onboarding_data')

    # If editing existing BOM
    bom = None
    if bom_id:
        bom = get_object_or_404(BillOfMaterial, pk=bom_id, product__company=company)

    products = Product.objects.filter(company=company)
    machines = Machine.objects.filter(company=company)
    existing_stages = ProductionStage.objects.all().values('name').distinct() # Suggest existing stage names

    context = {
        'products': products,
        'machines': machines,
        'existing_stages': existing_stages,
        'bom': bom,
    }
    return render(request, 'manufacturing/bom_builder.html', context)

@login_required
def api_save_bom(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)
        
    try:
        data = json.loads(request.body)
        user = request.user
        company = get_user_company(user)
        
        # 1. Parse Basic BOM Info
        product_id = data.get('product_id')
        base_qty = float(data.get('base_qty', 1))
        bom_status = data.get('status', 'draft')

        product = get_object_or_404(Product, pk=product_id, company=company)
        
        # Creating or Updating BOM
        bom_id = data.get('bom_id')
        if bom_id:
            bom = get_object_or_404(BillOfMaterial, pk=bom_id)
            if bom.status == 'active':
                 # In real app, we might force versioning here. For now, allow draft edit or error.
                 if bom_status == 'active': # User trying to save active as active?
                     pass
                 else:
                     return JsonResponse({"status": "error", "message": "Cannot edit Active BOM directly. Create new version."}, status=400)
            bom.base_quantity = base_qty
            bom.status = bom_status
            bom.save()
        else:
            bom = BillOfMaterial.objects.create(
                product=product,
                base_quantity=base_qty,
                status=bom_status,
                created_by=user
            )

        # 2. Process Materials (Tab 1)
        # Strategy: Delete existing and recreate (simplest for complex generic forms) 
        # OR update smart. For now: Wipe & Recreate for Drafts.
        bom.components.all().delete()
        for comp in data.get('components', []):
            BOMComponent.objects.create(
                bom=bom,
                material_name=comp['name'],
                quantity=comp['qty'],
                unit=comp['unit'],
                cost_per_unit=comp['cost'],
                wastage_quantity=comp.get('scrap_qty', 0),
                scrap_value_per_unit=comp.get('scrap_price', 0),
                scrap_type=comp.get('scrap_type', 'sell_as_scrap')
            )

        # 3. Process Operations (Tab 2)
        bom.operations.all().delete()
        for idx, op in enumerate(data.get('operations', [])):
            machine = None
            if op.get('machine_id'):
                machine = Machine.objects.get(pk=op['machine_id'])
            
            # Find or Create Stage
            stage_name = op.get('stage_name', f"Stage {idx+1}")
            stage, _ = ProductionStage.objects.get_or_create(
                name=stage_name,
                defaults={'machine': machine} 
                # Note: ProductionStage model links to generic machine preference, 
                # but BOMOperation links to specific machine for this BOM.
            )
            
            setup = normalize_operation_time_minutes(
                op.get('setup_time', 0),
                op.get('setup_time_unit') or op.get('setup_unit') or 'min',
            )
            run = normalize_operation_time_minutes(
                op.get('run_time', 0),
                op.get('run_time_unit') or op.get('run_unit') or 'min',
            )
            
            BOMOperation.objects.create(
                bom=bom,
                machine=machine,
                stage=stage,
                order=idx + 1,
                setup_time=setup,
                run_time=run,
                duration_minutes=max(int(round(float(setup + (run * base_qty)))), 0) # Backward compatibility
            )

        # 4. Process Acceptance Criteria (Tab 3)
        bom.acceptance_criteria.all().delete()
        for crit in data.get('criteria', []):
            BOMAcceptanceCriteria.objects.create(
                bom=bom,
                parameter=crit['parameter'],
                method=crit['method'],
                criteria_min=crit.get('min'),
                criteria_max=crit.get('max'),
                pass_fail=crit.get('pass_fail', False),
                target_value=crit.get('target_value'),
                tolerance=crit.get('tolerance'),
                is_critical=crit.get('is_critical', False)
            )

        return JsonResponse({"status": "success", "bom_id": bom.id, "message": "BOM Saved Successfully!"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

# 🧾 Create BOM (AJAX - Old - Deprecate later or keep for simple quick-adds)

@require_POST
def create_bom(request):
    """Create or Update a Bill of Materials."""
    try:
        bom_id = request.POST.get("bom_id")
        product_name = request.POST.get("product_name")
        version = request.POST.get("version", "v1.0").strip() # Strip whitespace
        status = request.POST.get("status", "draft")
        base_quantity = request.POST.get("base_quantity", 1)
        
        materials = request.POST.getlist("materials[]")
        quantities = request.POST.getlist("quantities[]")
        units = request.POST.getlist("units[]")
        costs = request.POST.getlist("costs[]")
        wastage = request.POST.getlist("wastage[]")
        scrap_values = request.POST.getlist("scrap_value[]")
        types = request.POST.getlist("types[]")
        scrap_types = request.POST.getlist("scrap_types[]")
        sub_bom_ids = request.POST.getlist("sub_bom_ids[]")

        if not product_name:
            return JsonResponse({"error": "Product Name is required"}, status=400)
        
        company = require_company(request.user)
        # Ensure product belongs to user's company
        product, _ = Product.objects.get_or_create(
            name=product_name,
            defaults={'company': company}
        )
        # If product exists but belongs to different company, create a new one
        if product.company != company:
            product = Product.objects.create(name=product_name, company=company)

        # 1. Create or Update BOM
        if bom_id:
            try:
                bom = BillOfMaterial.objects.filter(product__company=company).get(id=bom_id)
                
                # SMART UPDATE: If Active and Version Changed -> Create NEW (Fork)
                # Robust comparison: Strip whitespace to be safe
                # Only fork if status is ACTIVE. If Draft, allow version change in place.
                if bom.status == 'active' and str(bom.version).strip() != str(version).strip():
                    # Treat as New BOM (Fork)
                    bom = BillOfMaterial.objects.create(
                        product=product, 
                        version=version, 
                        status='draft', # FORCE DRAFT for new version to allow adding components
                        base_quantity=base_quantity,
                        created_by=request.user
                    )
                    # We do NOT delete old components since this is a new BOM
                else:
                    # Normal Update Logic
                    if bom.status == 'active': 
                         # Prevent editing active directly
                         return JsonResponse({"error": "Cannot modify an Active BOM directly. Change the Version number to save as a new version."}, status=400)
                    
                    bom.product = product
                    bom.version = version
                    bom.status = 'draft' # ⚡ Defer Activation: Keep draft while editing components
                    bom.base_quantity = base_quantity
                    bom.save()
                    
                    # Clear existing components/operations to replace them
                    bom.components.all().delete()
                    bom.operations.all().delete()
                
            except BillOfMaterial.DoesNotExist:
                return JsonResponse({"error": "BOM not found for update"}, status=404)
        else:
            bom = BillOfMaterial.objects.create(
                product=product, 
                version=version, 
                status='draft', # ⚡ Defer Activation
                base_quantity=base_quantity,
                created_by=request.user
            )

        # 2. Process Components
        for i, mat in enumerate(materials):
            mat = (mat or "").strip()
            if not mat:
                continue

            qty = quantities[i] if i < len(quantities) else 1
            unit = units[i] if i < len(units) else "pcs"
            cost = costs[i] if i < len(costs) else 0
            waste_qty = wastage[i] if i < len(wastage) else 0
            scrap_val = scrap_values[i] if i < len(scrap_values) else 0
            src_type = types[i] if i < len(types) else "raw"
            scr_type = scrap_types[i] if i < len(scrap_types) else "sell_as_scrap"
            sub_id = sub_bom_ids[i] if i < len(sub_bom_ids) and sub_bom_ids[i] else None
            
            sub_bom = None
            if sub_id:
                try:
                    # Ensure sub_bom belongs to same company
                    sub_bom = BillOfMaterial.objects.filter(product__company=company).get(id=sub_id)
                except: pass

            BOMComponent.objects.create(
                bom=bom,
                material_name=mat,
                quantity=qty or 1,
                unit=unit or "pcs",
                cost_per_unit=cost or 0,
                wastage_quantity=waste_qty or 0,
                scrap_value_per_unit=scrap_val or 0,
                source_type=src_type,
                scrap_type=scr_type,
                sub_bom=sub_bom
            )

        # 3. Process Operations (Routing)
        op_machines = request.POST.getlist("op_machines[]")
        op_stages = request.POST.getlist("op_stages[]")
        op_durations = request.POST.getlist("op_durations[]")

        for i, machine_id in enumerate(op_machines):
            # machine_id might be empty string if it's purely a stage? check logic
            # Earlier logic: if not machine_id: continue
            # But stages need to be saved too?
            # Revisit: toolbox gives machine_id for machine items.
            # Stage items: do they have machine_id?
            # JS addWorkflowNode: machine_id: type === 'machine' ? id : machineId
            # If stage, and no machineId, it's null.
            
            stage_id = op_stages[i] if i < len(op_stages) and op_stages[i] else None
            m_id = machine_id if machine_id else None
            
            # Allow saving if at least Stage OR Machine is present
            if not m_id and not stage_id:
                continue

            duration = op_durations[i] if i < len(op_durations) and op_durations[i] else 60

            BOMOperation.objects.create(
                bom=bom,
                machine_id=m_id,
                stage_id=stage_id,
                duration_minutes=int(duration),
                order=i+1
            )

        # 4. Final Activation (if requested)
        if status != 'draft':
            bom.status = status
            bom.save()

        return JsonResponse({"success": True})
    except Exception as e:
         import traceback
         traceback.print_exc()
         return JsonResponse({"error": str(e)}, status=500)


# 🔔 API: Get Notifications
@login_required
def api_notifications(request):
    """Fetch unread notifications for the user."""
    notifs = Notification.objects.filter(recipient=request.user, is_read=False).order_by('-created_at')
    data = [{
        "id": n.id,
        "title": n.title,
        "message": n.message,
        "link": n.link,
        "created_at": n.created_at.strftime("%Y-%m-%d %H:%M")
    } for n in notifs]
    return JsonResponse({"success": True, "notifications": data})

@login_required
def api_mark_read(request, notif_id):
    try:
        n = Notification.objects.get(id=notif_id, recipient=request.user)
        n.is_read = True
        n.save()
        return JsonResponse({"success": True})
    except Notification.DoesNotExist:
        return JsonResponse({"success": False, "error": "Not found"})


# 🧬 API: Get Work Order Materials (for Worker Input)
@login_required
def api_wo_materials(request, wo_id):
    try:
        wo = WorkOrder.objects.get(id=wo_id, company=require_company(request.user))
        materials = []
        if wo.bom:
            for comp in wo.bom.components.all():
                # Calculate expected Qty based on WO Quantity
                # Base Logic: (Comp Qty / Base Qty) * WO Qty
                ratio = (comp.quantity / wo.bom.base_quantity) if wo.bom.base_quantity else 0
                req_qty = ratio * wo.quantity
                
                materials.append({
                    "name": comp.material_name,
                    "qty": round(req_qty, 3),
                    "unit": comp.unit
                })
        
        return JsonResponse({"success": True, "materials": materials})
    except WorkOrder.DoesNotExist:
        return JsonResponse({"success": False, "error": "Work Order not found"})
    except Exception as e:
         import traceback
         traceback.print_exc()
         return JsonResponse({"error": str(e)}, status=500)


# ---------------------------------------------------------------------
# 📦 BOM Details (Modal View)
# ---------------------------------------------------------------------
@login_required
def view_bom_details(request, bom_id):
    """Return BOM details and total cost."""
    try:
        company = require_company(request.user)
        bom = BillOfMaterial.objects.filter(product__company=company).select_related("product", "created_by").prefetch_related("components").get(id=bom_id)
        total_cost = sum(c.total_cost() for c in bom.components.all())

        html = render_to_string("manufacturing/partials/bom_details.html", {
            "bom": bom,
            "components": bom.components.all(),
            "total_cost": total_cost,
        })
        return JsonResponse({"success": True, "html": html})
    except BillOfMaterial.DoesNotExist:
        return JsonResponse({"success": False, "error": "BOM not found"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ---------------------------------------------------------------------
# 🧰 Manage Production (Supervisor/Admin)
# ---------------------------------------------------------------------
@login_required
def manage_production(request):
    if not user_has_role(request.user, ['supervisor', 'admin']):
        return HttpResponseForbidden("🚫 You are not authorized to manage production.")
    
    company = require_company(request.user)
    workorders = WorkOrder.objects.filter(company=company).select_related('machine', 'bom').order_by('-id')
    pending_logs = ProductionLog.objects.filter(work_order__company=company, status='pending').select_related('worker', 'work_order')

    return render(request, 'manufacturing/manage_production.html', {
        "workorders": workorders,
        "pending_logs": pending_logs
    })


@require_POST
@login_required
def approve_log(request, log_id):
    """Approve or Reject a production log."""
    if not user_has_role(request.user, ['supervisor', 'admin']):
        return JsonResponse({"error": "Unauthorized"}, status=403)
    
    try:
        company = require_company(request.user)
        log = ProductionLog.objects.filter(work_order__company=company).get(id=log_id)
        action = request.POST.get("action") # 'approve' or 'reject'
        
        if action == 'approve':
            log.status = 'approved'
            log.reviewed_by = request.user
            log.save() 
            # Signal updates WorkOrder progress automatically
            return JsonResponse({"success": True, "message": "Log approved."})
        elif action == 'reject':
            log.status = 'rejected'
            log.reviewed_by = request.user
            log.save()
            return JsonResponse({"success": True, "message": "Log rejected."})
        
        return JsonResponse({"error": "Invalid action"}, status=400)

    except ProductionLog.DoesNotExist:
        return JsonResponse({"error": "Log not found"}, status=404)


# ---------------------------------------------------------------------
# 🧱 Record Output (Worker)
# ---------------------------------------------------------------------
@login_required
def record_output(request):
    if not user_has_role(request.user, ['worker']):
        return HttpResponseForbidden("🚫 You are not authorized to record output.")

    if request.method == "POST":
        form = ProductionLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.worker = request.user
            log.save()

            # 📉 Save Material Usage
            mat_names = request.POST.getlist('material_name[]')
            mat_qtys = request.POST.getlist('material_qty[]')
            mat_units = request.POST.getlist('material_unit[]')
            
            for i, name in enumerate(mat_names):
                qty = mat_qtys[i] if i < len(mat_qtys) and mat_qtys[i] else 0
                unit = mat_units[i] if i < len(mat_units) else 'pcs'
                
                try:
                    if float(qty) > 0:
                        MaterialUsage.objects.create(
                            production_log=log,
                            material_name=name,
                            quantity_used=qty,
                            unit=unit
                        )
                except: continue

            messages.success(request, "✅ Production log + Data submitted.")
            return redirect('record_output')
        else:
            messages.error(request, "❌ Error submitting log.")
    else:
        form = ProductionLogForm()

    company = get_user_company(request.user)
    # CRITICAL: Filter work orders to only show work orders from user's company
    if company:
        form.fields['work_order'].queryset = WorkOrder.objects.filter(company=company, status__in=['in_progress', 'pending'])
    
    # Show recent logs for this worker
    recent_logs = ProductionLog.objects.filter(worker=request.user, work_order__company=company).order_by('-created_at')[:10] if company else []

    return render(request, 'manufacturing/record_output.html', {"form": form, "recent_logs": recent_logs})

@login_required
def shop_floor(request):
    """
    Kiosk Mode View for Shop Floor Workers.
    Displays active job, queue, and quick actions.
    """
    # 1. Get User's Company
    company = get_user_company(request.user)
    if not company:
        return render(request, 'manufacturing/shop_floor.html', {"error": "No Company Assigned"})

    # 2. Get Active Work Order (In Progress)
    # Heuristic: The most recently updated 'in_progress' WO assigned to this user (or unassigned in their machine group)
    active_wo = WorkOrder.objects.filter(
        company=company, 
        status='in_progress', 
        assigned_to=request.user
    ).first()

    # 3. Get Queue (Pending)
    queue = WorkOrder.objects.filter(
        company=company,
        status='pending',
        assigned_to=request.user
    ).order_by('start_date')

    # 4. Metrics
    today = timezone.now().date()
    today_logs = ProductionLog.objects.filter(worker=request.user, created_at__date=today)
    today_output = sum([log.quantity for log in today_logs])

    context = {
        "active_wo": active_wo,
        "queue": queue,
        "today_output": today_output,
    }
    return render(request, 'manufacturing/shop_floor.html', context)


# ---------------------------------------------------------------------
# 🔧 Maintenance Dashboard
# ---------------------------------------------------------------------
@login_required
def maintenance_dashboard(request):
    """View machine statuses and handle faults."""
    if not user_has_role(request.user, ['maintenance', 'admin', 'supervisor']):
        return HttpResponseForbidden("🚫 You are not authorized to access maintenance.")

    company = require_company(request.user)
    
    if request.method == "POST":
        if not user_has_role(request.user, ['maintenance', 'admin']):
             return HttpResponseForbidden("🚫 Only Maintenance team can change status.")
        
        # Toggle machine status
        machine_id = request.POST.get("machine_id")
        new_status = request.POST.get("status")
        note = request.POST.get("note")
        
        try:
            machine = Machine.objects.filter(company=company).get(id=machine_id)
            if new_status:
                machine.status = new_status
            if note:
                machine.maintenance_note = note
            machine.save()
            
            # 🔔 Notify Planner if Broken
            if new_status in ['broken', 'maintenance']:
                Notification.objects.create(
                    recipient=company.created_by, # Notify Owner/Planner
                    title=f"Machine Alert: {machine.name}",
                    message=f"{machine.name} is now {new_status}. Action required.",
                    link="/manufacturing/planner/"
                )

            messages.success(request, f"✅ Updated {machine.name} status to {machine.get_status_display()}")
        except Machine.DoesNotExist:
            messages.error(request, "❌ Machine not found.")
        return redirect('maintenance_dashboard')

    machines = Machine.objects.filter(company=company).order_by('name')
    return render(request, 'manufacturing/maintenance_dashboard.html', {"machines": machines})


# ---------------------------------------------------------------------
# 🧪 Quality Check (Quality/Admin)
# ---------------------------------------------------------------------
from .models import QualityCheck

@login_required
def quality_check(request):
    if not user_has_role(request.user, ['quality', 'admin']):
        return HttpResponseForbidden("🚫 You are not authorized to perform quality checks.")
    
    if request.method == "POST":
        wo_id = request.POST.get("work_order")
        good = int(request.POST.get("good_quantity", 0))
        repair = int(request.POST.get("repair_quantity", 0))
        faulty = int(request.POST.get("faulty_quantity", 0))
        notes = request.POST.get("notes")

        try:
            company = require_company(request.user)
            wo = WorkOrder.objects.filter(company=company).get(id=wo_id)
            QualityCheck.objects.create(
                work_order=wo,
                checked_by=request.user,
                good_quantity=good,
                repair_quantity=repair,
                faulty_quantity=faulty,
                notes=notes
            )
            
            # Optional: Update WO status if fully checked?
            # For now just log it.
            messages.success(request, "✅ Quality check recorded.")
            return redirect('quality_check')
        except WorkOrder.DoesNotExist:
            messages.error(request, "❌ Work Order not found.")

    # List WOs that are in progress or completed and need checking
    company = require_company(request.user)
    work_orders = WorkOrder.objects.filter(company=company, status__in=['in_progress', 'completed']).order_by('-id')
    
    return render(request, 'manufacturing/quality_check.html', {"work_orders": work_orders})

# ---------------------------------------------------------------------
# 🔄 BOM Lifecycle Management (Master Rules)
# ---------------------------------------------------------------------
@require_POST
@login_required
def update_bom_status(request):
    """
    Handle BOM Lifecycle Transitions.
    DRAFT -> TEST -> ACTIVE -> ARCHIVED
    (Test mode is now optional - can go directly from DRAFT to ACTIVE)
    """
    try:
        bom_id = request.POST.get("bom_id")
        new_status = request.POST.get("status")

        if not bom_id or not new_status:
            return JsonResponse({"success": False, "error": "Missing parameters."})

        company = require_company(request.user)
        bom = BillOfMaterial.objects.filter(product__company=company).get(id=bom_id)
        current_status = bom.status

        # 1. DRAFT -> TEST
        if current_status == 'draft' and new_status == 'test':
            bom.status = 'test'
            bom.save()
            return JsonResponse({"success": True, "message": "Moved to Test Mode."})

        # 2. TEST -> DRAFT (Return to edit)
        elif current_status == 'test' and new_status == 'draft':
            bom.status = 'draft'
            bom.save()
            return JsonResponse({"success": True, "message": "Returned to Draft."})

        # 3. DRAFT -> ACTIVE (Direct activation - skip test mode)
        elif current_status == 'draft' and new_status == 'active':
            bom.status = 'active'
            bom.save()
            return JsonResponse({"success": True, "message": "BOM Activated directly from draft!"})

        # 4. TEST -> ACTIVE
        elif current_status == 'test' and new_status == 'active':
            bom.status = 'active'
            bom.save()
            return JsonResponse({"success": True, "message": "BOM Activated! Others archived."})

        # 5. ACTIVE -> ARCHIVED
        elif current_status == 'active' and new_status == 'archived':
            bom.status = 'archived'
            bom.save()
            return JsonResponse({"success": True, "message": "BOM Archived."})

        # 🚫 ILLEGAL TRANSITIONS
        else:
            return JsonResponse({
                "success": False, 
                "error": f"Illegal Transition: Cannot move from {current_status.upper()} to {new_status.upper()}."
            })

    except BillOfMaterial.DoesNotExist:
        return JsonResponse({"success": False, "error": "BOM not found."})
    except Exception as e:
        import traceback
        return JsonResponse({"success": False, "error": f"Server Error: {str(e)}", "trace": traceback.format_exc()})


# ---------------------------------------------------------------------
# 🧬 Get BOM JSON (For Simulation / Analysis)
# ---------------------------------------------------------------------
@login_required
def get_bom_json(request, bom_id):
    """Return full BOM structure for frontend simulation."""
    try:
        company = require_company(request.user)
        bom = BillOfMaterial.objects.filter(product__company=company).get(id=bom_id)
        
        components = []
        for c in bom.components.all():
            components.append({
                "material_name": c.material_name,
                "quantity": float(c.quantity),
                "unit": c.unit,
                "cost_per_unit": float(c.cost_per_unit),
                "wastage_quantity": float(c.wastage_quantity),
                "scrap_value_per_unit": float(c.scrap_value_per_unit),
                "source_type": c.source_type,
                "scrap_type": c.scrap_type,
                "sub_bom_id": c.sub_bom.id if c.sub_bom else None
            })

        operations = []
        for op in bom.operations.all().order_by('order'):
            operations.append({
                "id": op.id,
                "order": op.order,
                "duration": op.duration_minutes,
                "machine_id": op.machine.id if op.machine else None,
                "machine_name": op.machine.name if op.machine else None,
                "stage_id": op.stage.id if op.stage else None,
                "stage_name": op.stage.name if op.stage else None,
            })

        data = {
            "id": bom.id,
            "product_name": bom.product.name if bom.product else "Unknown",
            "version": bom.version,
            "status": bom.status,
            "base_quantity": float(bom.base_quantity),
            "uom": bom.uom,
            "uom": bom.uom,
            "components": components,
            "operations": operations
        }
        return JsonResponse({"success": True, "bom": data})

    except BillOfMaterial.DoesNotExist:
        return JsonResponse({"success": False, "error": "BOM not found."})
    except Exception as e:
        import traceback
        return JsonResponse({"success": False, "error": f"Server Error: {str(e)}", "trace": traceback.format_exc()})


# ---------------------------------------------------------------------
# 🏭 Supervisor Dashboard (Phase 3)
# ---------------------------------------------------------------------
@login_required
def supervisor_dashboard(request):
    """
    The Control Tower for the Supervisor.
    Shows Live Machine Grid and Pending Work Orders.
    """
    if not user_has_role(request.user, ['supervisor', 'admin', 'planner']): # Allow planner to see too
       return HttpResponseForbidden("Unauthorized")

    company = require_company(request.user)
    
    # Machines & Active WOs
    machines = Machine.objects.filter(company=company).order_by('name')
    pending_orders = WorkOrder.objects.filter(company=company, status__in=['pending', 'draft']).order_by('-start_date')
    pending_logs = ProductionLog.objects.filter(work_order__company=company, status='pending').select_related('worker', 'work_order')

    # Attach active WO manually to machine objects for template
    active_wos = {wo.machine_id: wo for wo in WorkOrder.objects.filter(company=company, status='in_progress', machine__isnull=False)}
    for m in machines:
        m.active_wo = active_wos.get(m.id)

    # HTMX Partial Rendering
    if request.GET.get('partial') == 'machine_grid':
        return render(request, 'manufacturing/partials/machine_grid.html', {
            'machine_list': machines
        })

    # Context items
    users = User.objects.filter(profile__company=company) # For Assignment
    
    return render(request, 'manufacturing/supervisor_dashboard.html', {
        'machine_list': machines,
        'pending_orders': pending_orders,
        'pending_logs': pending_logs,
        'pending_logs_count': pending_logs.count(),
        'users': users
    })





@require_POST
@login_required
def assign_work_order(request):
    """
    API to assign a Pending Work Order to a Machine and start it.
    """
    try:
        company = require_company(request.user)
        wo_id = request.POST.get('wo_id')
        machine_id = request.POST.get('machine_id')

        if not wo_id or not machine_id:
            return JsonResponse({'success': False, 'error': 'Missing parameters'})

        wo = WorkOrder.objects.filter(company=company).get(id=wo_id)
        machine = Machine.objects.filter(company=company).get(id=machine_id)

        # Update WO
        wo.machine = machine
        wo.status = 'in_progress' # Auto-start upon assignment? Or just 'assigned'? 
        # Let's say it moves to 'in_progress' to show up on the board.
        wo.save()

        # Update Machine Status (Optional automation)
        if machine.status == 'idle':
            machine.status = 'running'
            machine.save()

        return JsonResponse({'success': True, 'message': f'WO #{wo.id} assigned to {machine.name}'})

    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work Order not found'})
    except Machine.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Machine not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ---------------------------------------------------------------------
# 📡 Real-Time Shop Floor Control (Phase 3)
# ---------------------------------------------------------------------

@login_required
@require_POST
def log_production(request):
    """Handle production reporting from the shop floor."""
    try:
        data = request.POST
        wo_id = data.get('work_order_id')
        qty = int(data.get('quantity', 0))
        shift = data.get('shift', 'morning')
        note = data.get('note', '')

        company = require_company(request.user)
        work_order = WorkOrder.objects.filter(company=company).get(id=wo_id)
        
        # Create Log
        log = ProductionLog.objects.create(
            work_order=work_order,
            worker=request.user,
            quantity=qty,
            shift=shift,
            note=note,
            status='approved' # Auto-approve for now (or pending based on role)
        )
        
        # Signal 'update_work_order_progress' triggers automatically on save
        
        return JsonResponse({'success': True, 'message': 'Production logged successfully!'})
        
    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work Order not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})







@login_required
def api_wo_criteria(request, wo_id):
    """
    API: Fetch BOM Acceptance Criteria for a Work Order.
    """
    company = require_company(request.user)
    wo = get_object_or_404(WorkOrder, id=wo_id, company=company)
    
    criteria = []
    if wo.bom:
        for c in wo.bom.acceptance_criteria.all():
            criteria.append({
                'id': c.id,
                'parameter': c.parameter,
                'method': c.method,
                'min': c.criteria_min,
                'max': c.criteria_max,
                'target': c.target_value,
                'tolerance': c.tolerance,
                'pass_fail': c.pass_fail,
                'is_critical': c.is_critical
            })
            
    return JsonResponse({'success': True, 'criteria': criteria})


@login_required
@require_POST
def api_update_work_order_status(request, wo_id):
    """
    API: Update WO Status (Start, Pause, Complete)
    Called by Shop Floor Kiosk.
    """
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        company = require_company(request.user)
        
        wo = get_object_or_404(WorkOrder, id=wo_id, company=company)
        
        # Validation Logic Could Go Here (e.g. check transistions)
        
        if new_status:
            wo.status = new_status
            
            # Auto-assign if starting and not assigned
            if new_status == 'in_progress' and not wo.assigned_to:
                wo.assigned_to = request.user
            
            # Simple Progress Logic
            if new_status == 'completed':
                wo.progress = 100
                wo.actual_end = timezone.now()
            elif new_status == 'in_progress' and not wo.actual_start:
                wo.actual_start = timezone.now()
                
            wo.save()
            
            return JsonResponse({'status': 'success', 'message': f'Order #{wo.id} updated to {new_status}'})
            
        return JsonResponse({'status': 'error', 'message': 'No status provided'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



@login_required
def get_timeline_data(request):
    """API for Gantt Chart & Dashboard Polling"""
    from django.core.serializers.json import DjangoJSONEncoder
    from .services import DashboardService
    
    company = require_company(request.user)
    data = DashboardService.get_timeline_data(company)

    return JsonResponse({
        "success": True,
        **data
    }, encoder=DjangoJSONEncoder)


@login_required
def get_work_order(request, wo_id):
    """API to get work order details for editing"""
    try:
        company = require_company(request.user)
        work_order = WorkOrder.objects.get(id=wo_id, company=company)
        
        data = {
            'success': True,
            'work_order': {
                'id': work_order.id,
                'product_name': work_order.product_name,
                'quantity': work_order.quantity,
                'status': work_order.status,
                'priority': work_order.priority,
                'progress': float(work_order.progress),
                'start_date': work_order.start_date.isoformat() if work_order.start_date else None,
                'end_date': work_order.end_date.isoformat() if work_order.end_date else None,
                'machine_id': work_order.machine.id if work_order.machine else None,
                'machine_name': work_order.machine.name if work_order.machine else None,
                'assigned_to_id': work_order.assigned_to.id if work_order.assigned_to else None,
                'assigned_to_name': work_order.assigned_to.username if work_order.assigned_to else None,
            }
        }
        return JsonResponse(data)
    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work order not found'})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def update_work_order(request, wo_id):
    """API to update work order details"""
    try:
        company = require_company(request.user)
        work_order = WorkOrder.objects.get(id=wo_id, company=company)
        
        # Update fields
        work_order.product_name = request.POST.get('product_name', work_order.product_name)
        work_order.quantity = request.POST.get('quantity', work_order.quantity)
        work_order.status = request.POST.get('status', work_order.status)
        work_order.priority = request.POST.get('priority', work_order.priority)
        progress_value = request.POST.get('progress', work_order.progress)
        work_order.progress = progress_value if progress_value else 0
        
        # Handle dates
        start_date = request.POST.get('start_date')
        if start_date:
            from datetime import datetime
            work_order.start_date = datetime.fromisoformat(start_date.replace('T', ' '))
        
        end_date = request.POST.get('end_date')
        if end_date:
            from datetime import datetime
            work_order.end_date = datetime.fromisoformat(end_date.replace('T', ' '))
        
        work_order.save()
        
        return JsonResponse({'success': True, 'message': 'Work order updated successfully'})
    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work order not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ---------------------------------------------------------
# 📤 BULK IMPORT MODULE
# ---------------------------------------------------------
from .forms import BulkImportForm
import openpyxl
from django.contrib import messages

@login_required
def bulk_import_dashboard(request):
    if not user_has_role(request.user, ['Planner', 'Supervisor', 'Admin']):
        return HttpResponseForbidden("Access Denied")
        
    form = BulkImportForm()
    return render(request, 'manufacturing/bulk_import.html', {'form': form})

@login_required
def download_template(request, filename):
    """Serve Excel templates for bulk import."""
    import os
    from django.http import HttpResponse, Http404
    from django.conf import settings
    
    # Validation to prevent traversal
    allowed_files = [
        'machines_template.xlsx', 
        'products_template.xlsx', 
        'employees_template.xlsx', 
        'bom_template.xlsx',
        'work_orders_template.xlsx'
    ]
    
    if filename not in allowed_files:
        raise Http404("Template not found")
        
    file_path = os.path.join(settings.BASE_DIR, 'templates', 'downloads', filename)
    
    if not os.path.exists(file_path):
        raise Http404(f"File not found on server: {file_path}")
        
    with open(file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

@login_required
def handle_bulk_import(request):
    if request.method == 'POST':
        # Validate company assignment first
        try:
            company = require_company(request.user)
        except ValueError:
            messages.error(request, "❌ No company assigned to your account. Please contact administrator.")
            return redirect('bulk_import_dashboard')
        
        form = BulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            import_type = form.cleaned_data['import_type']
            file = request.FILES['file']
            
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                count = 0
                errors = []
                
                # Products Import
                if import_type == 'products':
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        # Row: Name | Type | Unit | Description | Cost
                        if row[0]: # Name required
                            product_name = str(row[0]).strip()
                            try:
                                # Try to get existing product
                                p = Product.objects.filter(name=product_name).first()
                                
                                if p:
                                    # Product exists - ensure it belongs to this company
                                    if p.company != company:
                                        errors.append(f"Row {row_idx}: Product '{product_name}' already exists but belongs to another company. Skipped.")
                                        continue
                                    # Update existing product
                                    p.material_type = row[1] or p.material_type or 'raw'
                                    p.unit = row[2] or p.unit or 'pcs'
                                    p.description = row[3] or p.description or ''
                                    p.company = company  # Ensure company is set
                                    p.save()
                                else:
                                    # Create new product
                                    p = Product.objects.create(
                                        name=product_name,
                                        company=company,
                                        material_type=row[1] or 'raw',
                                        unit=row[2] or 'pcs',
                                        description=row[3] or ''
                                    )
                                count += 1
                            except Exception as e:
                                errors.append(f"Row {row_idx}: Error importing '{product_name}': {str(e)}")
                                continue
                    
                    if errors:
                        messages.warning(request, f"✅ Imported {count} products. {len(errors)} errors occurred. Check console for details.")
                        for err in errors[:5]:  # Show first 5 errors
                            messages.warning(request, err)
                    else:
                        messages.success(request, f"✅ Successfully imported {count} products.")

                # Machines Import
                elif import_type == 'machines':
                    company = require_company(request.user)
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Row: Name | Code | Status | Type
                        if row[0]: # Name
                            Machine.objects.update_or_create(
                                company=company,
                                code=row[1] or f'M{count}',
                                defaults={
                                    'name': row[0],
                                    'status': row[2] or 'operational',
                                    'type': row[3] or 'General'
                                }
                            )
                            count += 1

                # Employees Import (New for Onboarding Step 3)
                elif import_type == 'employees':
                     from accounts.models import Role, Profile
                     # Row: Name | Email | Role | Department
                     for row in sheet.iter_rows(min_row=2, values_only=True):
                         name, email, role_name, dept = row[0], row[1], row[2], row[3]
                         if email and name:
                             if not User.objects.filter(username=email).exists():
                                 u = User.objects.create_user(username=email, email=email, password="Password123!", first_name=name)
                                 r, _ = Role.objects.get_or_create(name=role_name or 'worker')
                                 if hasattr(u, 'profile'):
                                     u.profile.role = r
                                     u.profile.company = request.user.profile.company
                                     u.profile.save()
                                 else:
                                     Profile.objects.create(user=u, role=r, company=request.user.profile.company)
                                 count += 1
                
                # BOM Import (Complex)
                elif import_type == 'bom':
                    # Flat format. Logic: Finds/Creates Product, Component, Operation
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        prod_name, comp_name, qty, op_name, duration, mach_code = row[0], row[1], row[2], row[3], row[4], row[5]
                        
                        if not prod_name: continue
                        
                        try:
                            prod_name = str(prod_name).strip()
                            # Try to get existing product from this company
                            product = Product.objects.filter(name=prod_name, company=company).first()
                            
                            if not product:
                                # Try to get product with same name (might belong to different company)
                                existing = Product.objects.filter(name=prod_name).first()
                                if existing and existing.company != company:
                                    errors.append(f"Row {row_idx}: Product '{prod_name}' exists but belongs to another company. Skipped.")
                                    continue
                                # Create new product for this company
                                product = Product.objects.create(
                                    name=prod_name,
                                    company=company
                                )
                            
                            # Ensure product belongs to company (safety check)
                            if product.company != company:
                                product.company = company
                                product.save()
                            
                            bom, created = BillOfMaterial.objects.get_or_create(
                                product=product, 
                                version="v1.0",
                                defaults={'status': 'draft', 'created_by': request.user}
                            )
                            
                            if comp_name:
                                 BOMComponent.objects.create(
                                     bom=bom,
                                     material_name=comp_name,
                                     quantity=qty or 1
                                 )
                             
                            if op_name:
                                machine = None
                                if mach_code:
                                    machine = Machine.objects.filter(company=company, code=mach_code).first()
                                stage, _ = ProductionStage.objects.get_or_create(
                                    name=op_name,
                                    machine=machine,
                                    defaults={'order': 1}
                                )
                                BOMOperation.objects.create(
                                    bom=bom,
                                    machine=machine,
                                    stage=stage,
                                    duration_minutes=int(duration or 60),
                                    order=bom.operations.count() + 1
                                )
                            count += 1
                        except Exception as e:
                            errors.append(f"Row {row_idx}: Error importing BOM for '{prod_name}': {str(e)}")
                            continue
                    
                    if errors:
                        messages.warning(request, f"✅ Imported {count} BOMs. {len(errors)} errors occurred.")
                        for err in errors[:5]:
                            messages.warning(request, err)
                    else:
                        messages.success(request, f"✅ Successfully imported {count} BOMs.")

                # Work Orders Import (New)
                elif import_type == 'work_orders':
                    company = require_company(request.user)
                    # Row: Product Name | Qty | Start | End | Status | Assigned Email
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                         prod, qty, start, end, status, email, mach_code = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
                         if prod:
                             user = None
                             machine = None
                             if email:
                                 user = User.objects.filter(profile__company=company, email=email).first()
                             
                             if mach_code:
                                machine = Machine.objects.filter(company=company, code=mach_code).first()

                             WorkOrder.objects.create(
                                 company=company,
                                 product_name=prod,
                                 quantity=qty or 1,
                                 start_date=start,
                                 end_date=end,
                                 status=status or 'pending',
                                 assigned_to=user,
                                 machine=machine
                             )
                             count += 1

                # Success message is handled per import type above
                if import_type not in ['products', 'machines', 'employees', 'bom']:
                    messages.success(request, f"✅ Successfully imported {count} items.")
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f"❌ Import Failed: {str(e)}")
                if hasattr(e, '__cause__') and e.__cause__:
                    messages.error(request, f"Details: {str(e.__cause__)}")
                
    # Redirect back to where they came from? Or always dashboard?
    # Referer check is brittle. Let's check query param 'next'
    next_url = request.POST.get('next', 'bulk_import_dashboard')
    return redirect(next_url)


# ---------------------------------------------------------
# 📊 REPORTING MODULE
# ---------------------------------------------------------
import csv
import json
from django.utils.dateparse import parse_date

@login_required
def reports_dashboard(request):
    """
    View daily production stats and list logs for printing.
    """
    if not user_has_role(request.user, ['Planner', 'Supervisor', 'Admin']):
        return HttpResponseForbidden("Access Denied")

    # Filter
    date_str = request.GET.get('date')
    if date_str:
        selected_date = parse_date(date_str) or timezone.now().date()
    else:
        selected_date = timezone.now().date()

    # Data
    company = require_company(request.user)
    logs = ProductionLog.objects.filter(work_order__company=company, date=selected_date).select_related('work_order', 'worker', 'work_order__bom')
    
    # Aggregates
    total_qty = logs.aggregate(Sum('quantity'))['quantity__sum'] or 0
    rejected_qty = logs.filter(status='rejected').aggregate(Sum('quantity'))['quantity__sum'] or 0
    approved_qty = logs.filter(status='approved').aggregate(Sum('quantity'))['quantity__sum'] or 0
    
    # Yield Rate
    if total_qty > 0:
        yield_rate = round((approved_qty / total_qty) * 100, 1)
    else:
        yield_rate = 100 # Default to 100 if no production yet
        
    # Costing Estimation (Simplified: Based on BOM Cost)
    estimated_cost = 0
    scrap_cost = 0
    
    for log in logs:
        # If WO has a BOM, use its cost. Otherwise assume 0 or dummy.
        if log.work_order.bom:
            unit_cost = log.work_order.bom.total_cost
            if log.status == 'approved':
                estimated_cost += float(unit_cost) * log.quantity
            elif log.status == 'rejected':
                scrap_cost += float(unit_cost) * log.quantity

    # 📊 Chart Data 1: Output by Shift (for selected date)
    shift_data = logs.values('shift').annotate(total=Sum('quantity')).order_by('shift')
    # predefined order
    shifts = ['morning', 'evening', 'night']
    shift_counts = [0, 0, 0]
    for item in shift_data:
        if item['shift'] in shifts:
            shift_counts[shifts.index(item['shift'])] = item['total']
    
    # 📈 Chart Data 2: 7-Day Output Trend
    # Get last 7 days including today
    today = timezone.now().date()
    last_7_days = [(today - timezone.timedelta(days=i)) for i in range(6, -1, -1)]
    trend_labels = [day.strftime('%b %d') for day in last_7_days]
    trend_data = []
    
    for day in last_7_days:
        daily_qty = ProductionLog.objects.filter(
            work_order__company=company,
            date=day, 
            status='approved'
        ).aggregate(Sum('quantity'))['quantity__sum'] or 0
        trend_data.append(daily_qty)

    context = {
        'selected_date': selected_date,
        'logs': logs,
        'total_qty': total_qty,
        'yield_rate': yield_rate,
        'rejected_qty': rejected_qty,
        'estimated_cost': round(estimated_cost, 2),
        'scrap_cost': round(scrap_cost, 2),
        # Chart Data (Raw Lists for json_script)
        'chart_shift_labels': ['Morning', 'Evening', 'Night'],
        'chart_shift_data': shift_counts,
        'chart_trend_labels': trend_labels,
        'chart_trend_data': trend_data,
        'chart_quality_data': [approved_qty, rejected_qty]
    }
    return render(request, 'manufacturing/reports_dashboard.html', context)


@login_required
def export_production_csv(request):
    """
    Export filtered production logs to CSV.
    """
    if not user_has_role(request.user, ['Planner', 'Supervisor', 'Admin']):
        return HttpResponseForbidden("Access Denied")

    date_str = request.GET.get('date')
    if date_str:
        selected_date = parse_date(date_str) or timezone.now().date()
    else:
        selected_date = timezone.now().date()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="production_report_{selected_date}.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Date', 'Shift', 'Product', 'Worker', 'Quantity', 'Status', 'Note'])

    company = require_company(request.user)
    logs = ProductionLog.objects.filter(work_order__company=company, date=selected_date).select_related('work_order', 'worker')
    for log in logs:
        writer.writerow([
            log.id, 
            log.date, 
            log.get_shift_display(), 
            log.work_order.product_name, 
            log.worker.username if log.worker else 'N/A', 
            log.quantity, 
            log.status,
            log.note
        ])

    return response
